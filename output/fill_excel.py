import fitz, re, sys, io
import openpyxl
from datetime import datetime
from copy import copy

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# Parse PDF
doc = fitz.open("input/חשבוניות ספק/bavli 21.25/תנועה אחרון/חשבוניות תנועה מעודכן.pdf")

DESC_MAP = {
    "שיפוצים עבודות עבור": "עבור עבודות שיפוצים",
    "בניה עבודות עבור": "עבור עבודות בניה",
}

invoices = []
for p in range(len(doc)):
    text = doc[p].get_text()

    m_inv = re.search(r"(\d{4,6})\s*מס חשבונית", text)
    inv_num = m_inv.group(1) if m_inv else ""

    m_date = re.search(r"(\d{2}/\d{2}/\d{4})\s*(?:עד|תאריך)", text)
    date_str = m_date.group(1) if m_date else ""

    m_totals = re.findall(r'סה₪([\d,]+\.\d{2})', text)
    amount_excl = float(m_totals[0].replace(",", "")) if m_totals else 0

    m_incl = re.search(r'לתשלום.*?סה₪([\d,]+\.\d{2})', text)
    amount_incl = float(m_incl.group(1).replace(",", "")) if m_incl else 0

    desc = ""
    for line in text.split("\n"):
        if line.startswith("1") and "₪" in line and "עבור" in line:
            d = re.search(r"^1(.+?)₪", line)
            if d:
                raw = d.group(1).strip()
                desc = DESC_MAP.get(raw, raw)
                break

    invoices.append({
        "page": p + 1,
        "inv_num": inv_num,
        "date": date_str,
        "amount_excl": amount_excl,
        "amount_incl": amount_incl,
        "desc": desc,
    })

doc.close()
print(f"Parsed {len(invoices)} invoices from PDF")

# Load Excel and delete existing data rows (rows 3-4, keep header rows 1-2)
xlsx_path = "input/חשבוניות ספק/bavli 21.25/תנועה אחרון/חשבוניות ספק תנועה מעודכן.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

# Delete data rows (row 3 onwards) - keep rows 1 (empty) and 2 (header)
# Also delete the summary rows at bottom
max_row = ws.max_row
if max_row > 2:
    ws.delete_rows(3, max_row - 2)
    print(f"Deleted rows 3-{max_row}")

# Copy style from header row for reference
header_row = 2

# Fill in invoice data starting at row 3
pdf_filename = "חשבוניות תנועה מעודכן"
for idx, inv in enumerate(invoices):
    row = idx + 3  # Start at row 3
    num = idx + 1

    # Parse date
    parts = inv["date"].split("/")
    if len(parts) == 3:
        dt = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
    else:
        dt = None

    # Col B (2): מס = sequential number
    ws.cell(row=row, column=2, value=num)
    # Col C (3): שם קובץ חשבונית
    ws.cell(row=row, column=3, value=pdf_filename)
    # Col D (4): מספר עמוד
    ws.cell(row=row, column=4, value=inv["page"])
    # Col E (5): חשבונית שנוצרה (empty for now)
    # Col F (6): מספר ספק
    ws.cell(row=row, column=6, value=60471)
    # Col G (7): תאריך חשבונית
    if dt:
        ws.cell(row=row, column=7, value=dt)
        ws.cell(row=row, column=7).number_format = "YYYY-MM-DD"
    # Col H (8): מספר חשבונית (BOOKNUM = invoice number)
    ws.cell(row=row, column=8, value=inv["inv_num"])
    # Col I (9): סניף
    ws.cell(row=row, column=9, value="015")
    # Col J (10): פרטים
    ws.cell(row=row, column=10, value=inv["desc"])
    # Col K (11): מספר הקצאה (not in this PDF)
    # Col L (12): מקט
    ws.cell(row=row, column=12, value="000")
    # Col M (13): תאור המוצר
    ws.cell(row=row, column=13, value=inv["desc"])
    # Col N (14): חשבון הוצאות
    ws.cell(row=row, column=14, value="2112-015")
    # Col O (15): סכום לפני מעמ
    ws.cell(row=row, column=15, value=inv["amount_excl"])
    # Col P (16): סכום כולל מע"מ
    ws.cell(row=row, column=16, value=inv["amount_incl"])

# Add summary row
summary_row = len(invoices) + 3
total_excl = sum(inv["amount_excl"] for inv in invoices)
total_incl = sum(inv["amount_incl"] for inv in invoices)
ws.cell(row=summary_row, column=15, value=total_excl)
ws.cell(row=summary_row, column=16, value=total_incl)

# VAT row
vat_row = summary_row + 2
ws.cell(row=vat_row, column=16, value=total_incl - total_excl)

wb.save(xlsx_path)
print(f"Saved {len(invoices)} invoices to Excel")
print(f"Total excl VAT: {total_excl:,.2f}")
print(f"Total incl VAT: {total_incl:,.2f}")
print(f"VAT: {total_incl - total_excl:,.2f}")
