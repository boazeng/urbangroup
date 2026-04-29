"""
Backend API server for Urban Group portal.
Proxies Priority ERP API calls, keeping credentials server-side.
"""

import sys
import os
import io
import json
import tempfile
import importlib.util
import logging
from pathlib import Path
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
    _IL_TZ = ZoneInfo("Asia/Jerusalem")
except ImportError:
    from datetime import timezone, timedelta
    _IL_TZ = timezone(timedelta(hours=2))


def _now_il():
    """Return current datetime in Israel timezone."""
    return datetime.now(_IL_TZ)

# Use logging (writes to stderr, works in Lambda where stdout is broken)
logger = logging.getLogger("urbangroup")
logger.setLevel(logging.INFO)
if not logger.handlers:
    handler = logging.StreamHandler(sys.stderr)
    handler.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(handler)

import requests as http_requests
from requests.auth import HTTPBasicAuth
import openpyxl
from flask import Flask, jsonify, send_file, request
from flask_cors import CORS

IS_LAMBDA = os.environ.get("IS_LAMBDA") == "true"

# In Lambda, code is deployed flat in lambda-backend/; locally, backend/ is one level down
if IS_LAMBDA:
    PROJECT_ROOT = Path(__file__).resolve().parent
else:
    PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Load agent 100 module
agent_100_path = PROJECT_ROOT / "agents" / "specific-mission-agents" / "priority-specific-agents" / "100-customer" / "100-customer_reader.py"
spec_100 = importlib.util.spec_from_file_location("customer_reader", agent_100_path)
customer_reader = importlib.util.module_from_spec(spec_100)
sys.modules["customer_reader"] = customer_reader
spec_100.loader.exec_module(customer_reader)

# Agents 200/210 wrap sys.stdout with UTF-8 at import time.
# Save a ref before each load to prevent GC from closing the underlying buffer.
_saved_stdout = sys.stdout

# Load agent 200 module (invoice writer)
agent_200_path = PROJECT_ROOT / "agents" / "specific-mission-agents" / "priority-specific-agents" / "200-invoices" / "200-invoice_writer.py"
spec_200 = importlib.util.spec_from_file_location("invoice_writer", agent_200_path)
invoice_writer = importlib.util.module_from_spec(spec_200)
sys.modules["invoice_writer"] = invoice_writer
spec_200.loader.exec_module(invoice_writer)

# Load agent 210 module (invoice closer)
agent_210_path = PROJECT_ROOT / "agents" / "specific-mission-agents" / "priority-specific-agents" / "210-invoice-closer" / "210-invoice_closer.py"
spec_210 = importlib.util.spec_from_file_location("invoice_closer", agent_210_path)
invoice_closer = importlib.util.module_from_spec(spec_210)
sys.modules["invoice_closer"] = invoice_closer
spec_210.loader.exec_module(invoice_closer)

# Load agent 300 module (service call writer)
agent_300_path = PROJECT_ROOT / "agents" / "specific-mission-agents" / "priority-specific-agents" / "300-service-call" / "300-service_call_writer.py"
spec_300 = importlib.util.spec_from_file_location("service_call_writer", agent_300_path)
service_call_writer = importlib.util.module_from_spec(spec_300)
sys.modules["service_call_writer"] = service_call_writer
spec_300.loader.exec_module(service_call_writer)

# Load aging report module
aging_report_path = PROJECT_ROOT / "agents" / "specific-mission-agents" / "priority-specific-agents" / "800-reports" / "aging_report.py"
spec_aging = importlib.util.spec_from_file_location("aging_report", aging_report_path)
aging_report = importlib.util.module_from_spec(spec_aging)
sys.modules["aging_report"] = aging_report
spec_aging.loader.exec_module(aging_report)

# Load AR1000 module (Ariel debt customer report)
ar1000_path = PROJECT_ROOT / "agents" / "flows" / "ariel" / "AR1000-debt-customer-report.py"
spec_ar1000 = importlib.util.spec_from_file_location("ar1000_report", ar1000_path)
ar1000_report = importlib.util.module_from_spec(spec_ar1000)
sys.modules["ar1000_report"] = ar1000_report
spec_ar1000.loader.exec_module(ar1000_report)

# Load AR10010 module (Ariel uncharged delivery notes)
ar10010_path = PROJECT_ROOT / "agents" / "flows" / "ariel" / "AR10010-uncharged-delivery.py"
spec_ar10010 = importlib.util.spec_from_file_location("ar10010_report", ar10010_path)
ar10010_report = importlib.util.module_from_spec(spec_ar10010)
sys.modules["ar10010_report"] = ar10010_report
spec_ar10010.loader.exec_module(ar10010_report)

# Load AR10020 module (Ariel consolidated invoices)
ar10020_path = PROJECT_ROOT / "agents" / "flows" / "ariel" / "AR10020-invoices.py"
spec_ar10020 = importlib.util.spec_from_file_location("ar10020_report", ar10020_path)
ar10020_report = importlib.util.module_from_spec(spec_ar10020)
sys.modules["ar10020_report"] = ar10020_report
spec_ar10020.loader.exec_module(ar10020_report)

# Load agent 5000 module (WhatsApp bot)
agent_5000_path = PROJECT_ROOT / "agents" / "tools-connection" / "5000-whatsapp" / "5000-whatsapp_bot.py"
spec_5000 = importlib.util.spec_from_file_location("whatsapp_bot", agent_5000_path)
whatsapp_bot = importlib.util.module_from_spec(spec_5000)
sys.modules["whatsapp_bot"] = whatsapp_bot
spec_5000.loader.exec_module(whatsapp_bot)

# Load agent 5010 module (Ariel WhatsApp bot)
agent_5010_path = PROJECT_ROOT / "agents" / "tools-connection" / "5010-whatsapp" / "5010-whatsapp_bot.py"
spec_5010 = importlib.util.spec_from_file_location("whatsapp_bot_ariel", agent_5010_path)
whatsapp_bot_ariel = importlib.util.module_from_spec(spec_5010)
sys.modules["whatsapp_bot_ariel"] = whatsapp_bot_ariel
spec_5010.loader.exec_module(whatsapp_bot_ariel)

# Load ALLM1000 (Ariel command parser LLM)
allm1000_path = PROJECT_ROOT / "agents" / "LLM" / "ariel" / "ALLM1000-command-parser" / "ALLM1000_command_parser.py"
spec_allm1000 = importlib.util.spec_from_file_location("allm1000_command_parser", allm1000_path)
allm1000_module = importlib.util.module_from_spec(spec_allm1000)
sys.modules["allm1000_command_parser"] = allm1000_module
spec_allm1000.loader.exec_module(allm1000_module)

# Load PDF generator for Ariel reports
pdf_gen_path = PROJECT_ROOT / "agents" / "LLM" / "ariel" / "ALLM1000-command-parser" / "pdf_generator.py"
spec_pdf_gen = importlib.util.spec_from_file_location("pdf_generator", pdf_gen_path)
pdf_gen_module = importlib.util.module_from_spec(spec_pdf_gen)
sys.modules["pdf_generator"] = pdf_gen_module
spec_pdf_gen.loader.exec_module(pdf_gen_module)

# Load A1000 bot (Ariel WhatsApp smart bot)
a1000_path = PROJECT_ROOT / "agents" / "smart-agents-and-bots" / "ariel" / "A1000-ariel-whatsapp-bot" / "A1000_bot.py"
spec_a1000 = importlib.util.spec_from_file_location("a1000_bot", a1000_path)
a1000_bot = importlib.util.module_from_spec(spec_a1000)
sys.modules["a1000_bot"] = a1000_bot
spec_a1000.loader.exec_module(a1000_bot)

# Load M1000 bot (maintenance WhatsApp smart bot)
m1000_path = PROJECT_ROOT / "agents" / "smart-agents-and-bots" / "maintenance" / "M1000-maintenance-whatsapp-bot" / "M1000_bot.py"
spec_m1000 = importlib.util.spec_from_file_location("m1000_bot", m1000_path)
m1000_bot = importlib.util.module_from_spec(spec_m1000)
sys.modules["m1000_bot"] = m1000_bot
spec_m1000.loader.exec_module(m1000_bot)

# Load M10010 bot (troubleshooting script bot)
m10010_path = PROJECT_ROOT / "agents" / "smart-agents-and-bots" / "maintenance" / "M10010-troubleshoot-bot" / "M10010_bot.py"
spec_m10010 = importlib.util.spec_from_file_location("m10010_bot", m10010_path)
m10010_bot = importlib.util.module_from_spec(spec_m10010)
sys.modules["m10010_bot"] = m10010_bot
spec_m10010.loader.exec_module(m10010_bot)

# Load agent 420 module (invoice printer / attachment downloader)
agent_420_path = PROJECT_ROOT / "agents" / "specific-mission-agents" / "priority-specific-agents" / "420-invoice-printer" / "420-invoice_printer.py"
spec_420 = importlib.util.spec_from_file_location("invoice_printer", agent_420_path)
invoice_printer = importlib.util.module_from_spec(spec_420)
sys.modules["invoice_printer"] = invoice_printer
spec_420.loader.exec_module(invoice_printer)

# Load maintenance database module
maint_db_path = PROJECT_ROOT / "database" / "maintenance" / "maintenance_db.py"
spec_maint_db = importlib.util.spec_from_file_location("maintenance_db", maint_db_path)
maintenance_db = importlib.util.module_from_spec(spec_maint_db)
sys.modules["maintenance_db"] = maintenance_db
spec_maint_db.loader.exec_module(maintenance_db)

# Load troubleshoot sessions database module
ts_db_path = PROJECT_ROOT / "database" / "maintenance" / "troubleshoot_sessions_db.py"
spec_ts_db = importlib.util.spec_from_file_location("troubleshoot_sessions_db", ts_db_path)
troubleshoot_sessions_db = importlib.util.module_from_spec(spec_ts_db)
sys.modules["troubleshoot_sessions_db"] = troubleshoot_sessions_db
spec_ts_db.loader.exec_module(troubleshoot_sessions_db)

# Load bot scripts database module
bs_db_path = PROJECT_ROOT / "database" / "maintenance" / "bot_scripts_db.py"
spec_bs_db = importlib.util.spec_from_file_location("bot_scripts_db", bs_db_path)
bot_scripts_db = importlib.util.module_from_spec(spec_bs_db)
sys.modules["bot_scripts_db"] = bot_scripts_db
spec_bs_db.loader.exec_module(bot_scripts_db)

# Load bot prompts database module
bp_db_path = PROJECT_ROOT / "database" / "maintenance" / "bot_prompts_db.py"
spec_bp_db = importlib.util.spec_from_file_location("bot_prompts_db", bp_db_path)
bot_prompts_db = importlib.util.module_from_spec(spec_bp_db)
sys.modules["bot_prompts_db"] = bot_prompts_db
spec_bp_db.loader.exec_module(bot_prompts_db)

# Load delivery notes database module
dn_db_path = PROJECT_ROOT / "database" / "maintenance" / "delivery_notes_db.py"
spec_dn_db = importlib.util.spec_from_file_location("delivery_notes_db", dn_db_path)
delivery_notes_db = importlib.util.module_from_spec(spec_dn_db)
sys.modules["delivery_notes_db"] = delivery_notes_db
spec_dn_db.loader.exec_module(delivery_notes_db)

# Load knowledge database module
kn_db_path = PROJECT_ROOT / "database" / "maintenance" / "knowledge_db.py"
spec_kn_db = importlib.util.spec_from_file_location("knowledge_db", kn_db_path)
knowledge_db = importlib.util.module_from_spec(spec_kn_db)
sys.modules["knowledge_db"] = knowledge_db
spec_kn_db.loader.exec_module(knowledge_db)

# Load RAG retrieval module
rag_path = PROJECT_ROOT / "agents" / "LLM" / "maintenance" / "rag_retrieval.py"
spec_rag = importlib.util.spec_from_file_location("rag_retrieval", rag_path)
rag_retrieval = importlib.util.module_from_spec(spec_rag)
sys.modules["rag_retrieval"] = rag_retrieval
spec_rag.loader.exec_module(rag_retrieval)

# Load LLM2000 module (invoice analyzer)
llm2000_path = PROJECT_ROOT / "agents" / "LLM" / "LLM2000-invoice-analyzer" / "LLM2000_invoice_analyzer.py"
spec_llm2000 = importlib.util.spec_from_file_location("llm2000_invoice_analyzer", llm2000_path)
llm2000_analyzer = importlib.util.module_from_spec(spec_llm2000)
sys.modules["llm2000_invoice_analyzer"] = llm2000_analyzer
spec_llm2000.loader.exec_module(llm2000_analyzer)

# Ensure stdout is usable (use the latest UTF-8 wrapper or restore original)
if sys.stdout.closed:
    sys.stdout = _saved_stdout

app = Flask(__name__)
CORS(app)

# ── Seed default bot script on startup ───────────────────────
try:
    m10010_bot.seed_default_script()
except Exception as _seed_err:
    logger.warning(f"Bot script seed skipped: {_seed_err}")

# ── Seed default LLM prompt on startup ───────────────────
try:
    existing_prompt = bot_prompts_db.get_active_prompt(use_cache=False)
    if not existing_prompt:
        # Import the hardcoded prompt from MLLM1000
        _mllm_path = PROJECT_ROOT / "agents" / "LLM" / "maintenance" / "MLLM1000-servicecall-identifier" / "MLLM1000_servicecall_identifier.py"
        _spec_mllm = importlib.util.spec_from_file_location("_mllm_seed", _mllm_path)
        _mllm_mod = importlib.util.module_from_spec(_spec_mllm)
        _spec_mllm.loader.exec_module(_mllm_mod)
        bot_prompts_db.save_prompt({
            "prompt_id": "servicecall-identifier-v1",
            "name": "Service Call Identifier - Default",
            "content": _mllm_mod.SYSTEM_PROMPT,
            "active": True,
        })
        logger.info("Seeded default LLM prompt from MLLM1000")
except Exception as _prompt_seed_err:
    logger.warning(f"Bot prompt seed skipped: {_prompt_seed_err}")

# ── Environment switching (demo / real) ──────────────────────
# Fallback to PRIORITY_URL for backward compat (e.g. Lambda with old config)
_fallback_url = os.getenv("PRIORITY_URL", "")
PRIORITY_URL_DEMO = os.getenv("PRIORITY_URL_DEMO", _fallback_url).rstrip("/")
PRIORITY_URL_REAL = os.getenv("PRIORITY_URL_REAL", "").rstrip("/")


def get_priority_url():
    """Return the Priority URL matching the request's ?env= param."""
    env = request.args.get("env", "demo")
    return PRIORITY_URL_REAL if env == "real" else PRIORITY_URL_DEMO


def set_priority_env():
    """Set PRIORITY_URL on all agent modules according to the request env."""
    url = get_priority_url()
    customer_reader.PRIORITY_URL = url
    invoice_writer.PRIORITY_URL = url
    service_call_writer.PRIORITY_URL = url
    aging_report.PRIORITY_URL = url
    ar1000_report.PRIORITY_URL = url
    ar10010_report.PRIORITY_URL = url
    ar10020_report.PRIORITY_URL = url
    invoice_printer.PRIORITY_URL = url


@app.route("/api/customers", methods=["GET"])
def get_customers():
    """Fetch customer list from Priority ERP."""
    set_priority_env()
    try:
        customers = customer_reader.fetch_customers(top=9999)
        return jsonify({"ok": True, "customers": customers})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/template", methods=["GET"])
def download_template():
    """Download the invoice template Excel file."""
    if IS_LAMBDA:
        import boto3
        s3 = boto3.client("s3")
        bucket = os.environ["TEMPLATE_BUCKET"]
        key = os.environ["TEMPLATE_KEY"]
        tmp_path = "/tmp/template.xlsx"
        s3.download_file(bucket, key, tmp_path)
        return send_file(
            tmp_path,
            as_attachment=True,
            download_name="חשבונית עמלת גבייה - טמפלט.xlsx",
        )
    template_path = PROJECT_ROOT / "input" / "חשבונית עמלת גבייה - טמפלט.xlsx"
    if not template_path.exists():
        return jsonify({"ok": False, "error": "Template file not found"}), 404
    return send_file(
        template_path,
        as_attachment=True,
        download_name="חשבונית עמלת גבייה - טמפלט.xlsx",
    )


def parse_excel_invoices(filepath):
    """Parse uploaded Excel file into invoice data list."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    invoices = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        row_num = row[1].value       # B: מס
        date_val = row[2].value      # C: תאריך חשבונית
        details = row[3].value       # D: פרטים
        branch = row[4].value        # E: סניף
        custname = row[5].value      # F: מספר לקוח פריוריטי
        cust_label = row[6].value    # G: שם לקוח
        partname = row[7].value      # H: מקט
        part_desc = row[8].value     # I: תאור מוצר
        quantity = row[9].value      # J: כמות
        price_with_vat = row[11].value  # L: סכום כולל מעמ

        if not custname:
            continue

        # Format date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        elif date_val:
            date_str = str(date_val)
        else:
            date_str = _now_il().strftime("%Y-%m-%d")

        # Always calculate price from column L (with VAT)
        price_no_vat = round(price_with_vat / 1.18, 2) if price_with_vat else 0

        item = {
            "PARTNAME": str(partname).strip(),
            "TQUANT": quantity or 1,
            "PRICE": price_no_vat,
        }
        if part_desc:
            item["PDES"] = str(part_desc).strip()

        invoices.append({
            "row": row_num,
            "CUSTNAME": str(custname).strip(),
            "CUST_LABEL": str(cust_label or "").strip(),
            "IVDATE": date_str,
            "BRANCHNAME": str(branch or "000").strip(),
            "DETAILS": str(details or "").strip(),
            "items": [item],
        })

    wb.close()
    return invoices


@app.route("/api/invoices/run", methods=["POST"])
def run_invoices():
    """Process uploaded Excel file and create invoices in Priority."""
    set_priority_env()
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded"}), 400

    uploaded = request.files["file"]
    should_finalize = request.form.get("finalize", "1") == "1"

    # Save to temp file (Lambda only allows /tmp)
    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix=".xlsx", dir="/tmp" if IS_LAMBDA else None
    )
    uploaded.save(tmp.name)
    tmp.close()

    try:
        invoices = parse_excel_invoices(tmp.name)
    except Exception as e:
        os.unlink(tmp.name)
        return jsonify({"ok": False, "error": f"שגיאה בקריאת הקובץ: {e}"}), 400

    os.unlink(tmp.name)

    if not invoices:
        return jsonify({"ok": False, "error": "לא נמצאו שורות נתונים בקובץ"}), 400

    results = []
    for inv in invoices:
        try:
            result = invoice_writer.create_invoice(
                inv["CUSTNAME"],
                inv["IVDATE"],
                inv["BRANCHNAME"],
                inv["items"],
                details=inv.get("DETAILS"),
            )
            ivnum = result.get("IVNUM", "")

            # Finalize the invoice (טיוטא → סופית) if requested
            finalized = False
            finalize_error = None
            if should_finalize and ivnum:
                try:
                    invoice_closer.finalize_invoice(ivnum)
                    finalized = True
                except Exception as fe:
                    finalize_error = str(fe)

            results.append({
                "row": inv["row"],
                "customer": inv["CUSTNAME"],
                "name": inv["CUST_LABEL"],
                "status": "OK",
                "ivnum": ivnum or "N/A",
                "totprice": result.get("TOTPRICE", 0),
                "finalized": finalized,
                "finalize_error": finalize_error,
            })
        except Exception as e:
            results.append({
                "row": inv["row"],
                "customer": inv["CUSTNAME"],
                "name": inv["CUST_LABEL"],
                "status": "FAILED",
                "error": str(e),
            })

    success = sum(1 for r in results if r["status"] == "OK")
    failed = sum(1 for r in results if r["status"] == "FAILED")

    return jsonify({
        "ok": True,
        "total": len(results),
        "success": success,
        "failed": failed,
        "invoices": results,
    })


# ── WhatsApp Webhook Routes ──────────────────────────────────

@app.route("/api/whatsapp/webhook", methods=["GET"])
def whatsapp_verify():
    """Meta webhook verification (subscription handshake)."""
    mode = request.args.get("hub.mode", "")
    token = request.args.get("hub.verify_token", "")
    challenge = request.args.get("hub.challenge", "")

    result = whatsapp_bot.verify_webhook(mode, token, challenge)
    if result:
        return result, 200
    return "Forbidden", 403


@app.route("/api/whatsapp/webhook", methods=["POST"])
def whatsapp_incoming():
    """Receive incoming WhatsApp messages from Meta."""
    payload = request.get_json(silent=True) or {}
    messages = whatsapp_bot.handle_incoming(payload)
    logger.info(f"Webhook received: {len(messages)} message(s)")

    # Mark each message as read
    for msg in messages:
        if msg.get("message_id"):
            whatsapp_bot.mark_as_read(msg["message_id"])

    for msg in messages:
        phone = msg.get("phone", "")
        logger.info(f"From {phone} ({msg.get('name')}): {msg.get('text', '')[:100]}")

        try:
            # Reset keywords — clear session and start fresh
            RESET_KEYWORDS = {"התחל", "התחל מחדש", "חזור", "תפריט", "0", "reset", "restart", "menu"}
            text_stripped = (msg.get("text", "") or "").strip()
            if text_stripped in RESET_KEYWORDS:
                m10010_bot.reset_session(phone)
                logger.info(f"Session reset by keyword for {phone}")
                # Fall through to normal M1000 flow below

            # Check if this phone has an active troubleshooting session
            elif m10010_bot.get_active_session(phone):
                result = m10010_bot.process_message(
                    phone=phone,
                    text=msg.get("text", ""),
                    msg_type=msg.get("type", "text"),
                    caption=msg.get("caption", ""),
                )
                if result:
                    _send_bot_response(phone, result)
                    logger.info(f"M10010 reply sent to {phone}")
                continue

            # Normal M1000 flow
            response = m1000_bot.process_message(
                phone=phone,
                name=msg.get("name", ""),
                text=msg.get("text", ""),
                msg_type=msg.get("type", "text"),
                message_id=msg.get("message_id", ""),
                media_id=msg.get("media_id", ""),
                caption=msg.get("caption", ""),
            )

            if isinstance(response, dict) and response.get("voice_bot_handled"):
                # Voice bot: service call created directly, no reply needed
                logger.info(f"Voice bot call created for {phone}: "
                            f"DOCNO={response.get('priority_callno', '')} "
                            f"ID={response.get('call_id', '')}")
            elif isinstance(response, dict) and response.get("handoff") == "M10010":
                # M1000 hands off - start troubleshooting session
                result = m10010_bot.start_session(
                    phone=phone,
                    name=msg.get("name", ""),
                    llm_result=response.get("llm_result", {}),
                    parsed_data=response.get("parsed_data", {}),
                    message_id=msg.get("message_id", ""),
                    media_id=msg.get("media_id", ""),
                    original_text=response.get("original_text", msg.get("text", "")),
                    device_number=response.get("device_number", ""),
                    customer_number=response.get("customer_number", ""),
                    customer_name=response.get("customer_name", ""),
                    script_id=response.get("script_id"),
                )
                if result:
                    _send_bot_response(phone, result)
                    logger.info(f"M10010 session started for {phone}")
            elif response:
                whatsapp_bot.send_message(phone, response)
                logger.info(f"M1000 reply sent to {phone}")

        except Exception as e:
            logger.error(f"Bot error for {phone}: {e}")

    return jsonify({"ok": True}), 200


def _send_bot_response(phone, result):
    """Send a bot response - either buttons or plain text."""
    text = result.get("text", "")
    if text:  # Skip sending if text is empty (e.g. silent done actions)
        if result.get("buttons"):
            whatsapp_bot.send_buttons(
                phone,
                text,
                result["buttons"],
                header=result.get("header"),
                footer=result.get("footer"),
            )
        else:
            whatsapp_bot.send_message(phone, text)

    # Send admin notification if M10010 requested one (e.g. voice bot service call)
    notify = result.get("notify_whatsapp")
    if notify and notify.get("phone") and notify.get("text"):
        try:
            whatsapp_bot.send_message(notify["phone"], notify["text"])
            logger.info(f"Admin notification sent to {notify['phone']}")
        except Exception as e:
            logger.error(f"Admin notification failed: {e}")


@app.route("/api/whatsapp/send", methods=["POST"])
def whatsapp_send():
    """Send a WhatsApp message (from portal or other agents)."""
    data = request.get_json(silent=True) or {}
    phone = data.get("phone", "")
    text = data.get("text", "")
    template_name = data.get("template")

    if not phone:
        return jsonify({"ok": False, "error": "Missing phone number"}), 400

    try:
        if template_name:
            result = whatsapp_bot.send_template(
                phone,
                template_name,
                language=data.get("language", "he"),
                parameters=data.get("parameters"),
            )
        else:
            if not text:
                return jsonify({"ok": False, "error": "Missing text"}), 400
            result = whatsapp_bot.send_message(phone, text)

        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── WhatsApp Ariel Webhook Routes ────────────────────────────

@app.route("/api/whatsapp-ariel/webhook", methods=["GET"])
def whatsapp_ariel_verify():
    """Meta webhook verification for Ariel number."""
    mode = request.args.get("hub.mode", "")
    token = request.args.get("hub.verify_token", "")
    challenge = request.args.get("hub.challenge", "")

    result = whatsapp_bot_ariel.verify_webhook(mode, token, challenge)
    if result:
        return result, 200
    return "Forbidden", 403


@app.route("/api/whatsapp-ariel/webhook", methods=["POST"])
def whatsapp_ariel_incoming():
    """Receive incoming WhatsApp messages from Ariel Meta App."""
    payload = request.get_json(silent=True) or {}
    messages = whatsapp_bot_ariel.handle_incoming(payload)
    logger.info(f"Ariel webhook received: {len(messages)} message(s)")

    for msg in messages:
        if msg.get("message_id"):
            whatsapp_bot_ariel.mark_as_read(msg["message_id"])

    for msg in messages:
        logger.info(f"Ariel from {msg.get('phone')} ({msg.get('name')}): {msg.get('text', '')[:100]}")
        try:
            response = a1000_bot.process_message(
                phone=msg.get("phone", ""),
                name=msg.get("name", ""),
                text=msg.get("text", ""),
                msg_type=msg.get("type", "text"),
                message_id=msg.get("message_id", ""),
                media_id=msg.get("media_id", ""),
                caption=msg.get("caption", ""),
            )
            if response:
                whatsapp_bot_ariel.send_message(msg["phone"], response)
                logger.info(f"A1000 reply sent to {msg['phone']}")
        except Exception as e:
            logger.error(f"A1000 bot error: {e}")

    return jsonify({"ok": True}), 200


@app.route("/api/whatsapp-ariel/send", methods=["POST"])
def whatsapp_ariel_send():
    """Send a WhatsApp message from the Ariel number."""
    data = request.get_json(silent=True) or {}
    phone = data.get("phone", "")
    text = data.get("text", "")
    template_name = data.get("template")

    if not phone:
        return jsonify({"ok": False, "error": "Missing phone number"}), 400

    try:
        if template_name:
            result = whatsapp_bot_ariel.send_template(
                phone,
                template_name,
                language=data.get("language", "he"),
                parameters=data.get("parameters"),
            )
        else:
            if not text:
                return jsonify({"ok": False, "error": "Missing text"}), 400
            result = whatsapp_bot_ariel.send_message(phone, text)

        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Messages (DynamoDB) ──────────────────────────────────────

@app.route("/api/messages", methods=["GET"])
def get_messages():
    """Get WhatsApp messages from DynamoDB."""
    status = request.args.get("status")
    limit = int(request.args.get("limit", "50"))
    try:
        messages = maintenance_db.get_messages(status=status, limit=limit)
        return jsonify({"ok": True, "messages": messages, "count": len(messages)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/messages/<item_id>/status", methods=["PUT"])
def update_message_status(item_id):
    """Update a message status."""
    data = request.get_json(silent=True) or {}
    new_status = data.get("status")
    if not new_status:
        return jsonify({"ok": False, "error": "Missing status"}), 400
    try:
        updated = maintenance_db.update_message_status(item_id, new_status)
        return jsonify({"ok": True, "message": updated})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Service Calls (DynamoDB) ─────────────────────────────────

@app.route("/api/service-calls", methods=["GET"])
def get_service_calls():
    """Get service calls from DynamoDB."""
    status = request.args.get("status")
    phone = request.args.get("phone")
    limit = int(request.args.get("limit", "50"))
    try:
        calls = maintenance_db.get_service_calls(status=status, phone=phone, limit=limit)
        return jsonify({"ok": True, "service_calls": calls, "count": len(calls)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/service-calls/<item_id>/status", methods=["PUT"])
def update_service_call_status(item_id):
    """Update a service call status."""
    data = request.get_json(silent=True) or {}
    new_status = data.get("status")
    if not new_status:
        return jsonify({"ok": False, "error": "Missing status"}), 400
    try:
        updated = maintenance_db.update_service_call_status(item_id, new_status)
        return jsonify({"ok": True, "service_call": updated})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/service-calls/<item_id>/push", methods=["POST"])
def push_service_call_to_priority(item_id):
    """Push a service call to Priority ERP via Agent 300."""
    set_priority_env()
    try:
        call = maintenance_db.get_service_call(item_id)
        if not call:
            return jsonify({"ok": False, "error": "קריאת שירות לא נמצאה"}), 404

        if call.get("priority_pushed"):
            return jsonify({"ok": False, "error": "קריאת השירות כבר נשלחה לפריוריטי"}), 400

        result = service_call_writer.create_service_call(call)
        callno = str(result.get("DOCNO", ""))
        maintenance_db.mark_service_call_pushed(item_id, callno=callno)

        return jsonify({
            "ok": True,
            "callno": callno,
            "priority_response": result,
        })
    except Exception as e:
        logger.error(f"Error pushing service call {item_id}: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Reports ──────────────────────────────────────────────────

@app.route("/api/reports/aging", methods=["GET"])
def get_aging_report():
    """Generate aging report for consolidated invoices."""
    set_priority_env()
    branch = request.args.get("branch") or None
    try:
        report = aging_report.fetch_aging_report(branch=branch)
        return jsonify({"ok": True, **report})
    except Exception as e:
        logger.error(f"Error generating aging report: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/reports/ariel-debt", methods=["GET"])
def get_ariel_debt_report():
    """Generate AR1000 Ariel debt customer report."""
    set_priority_env()
    try:
        report = ar1000_report.generate_report()
        return jsonify({"ok": True, **report})
    except Exception as e:
        logger.error(f"Error generating Ariel debt report: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/reports/ariel-uncharged-delivery", methods=["GET"])
def get_ariel_uncharged_delivery():
    """Generate AR10010 uncharged delivery notes report."""
    set_priority_env()
    try:
        report = ar10010_report.generate_report()
        return jsonify({"ok": True, **report})
    except Exception as e:
        logger.error(f"Error generating uncharged delivery report: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/reports/ariel-invoices", methods=["GET"])
def get_ariel_invoices():
    """Generate AR10020 consolidated invoices report."""
    set_priority_env()
    days_back = int(request.args.get("days_back", "30"))
    try:
        report = ar10020_report.generate_report(days_back=days_back)
        return jsonify({"ok": True, **report})
    except Exception as e:
        logger.error(f"Error generating invoices report: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Customer Invoices (ניהול חשבוניות) ──────────────────────

@app.route("/api/hr/customer-invoices", methods=["GET"])
def get_customer_invoices():
    """Fetch last 10 CINVOICES for a customer (branch 102, finalized)."""
    try:
        customer = request.args.get("customer", "").strip()
        if customer.endswith("-102"):
            customer = customer[:-4]
        if not customer:
            return jsonify({"ok": False, "error": "Missing customer"}), 400

        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        headers = {"Accept": "application/json", "OData-Version": "4.0"}

        top = int(request.args.get("top", "10"))
        api_url = (
            f"{url}/CINVOICES"
            f"?$filter=BRANCHNAME eq '102' and CUSTNAME eq '{customer}' and FINAL eq 'Y'"
            f"&$select=IVNUM,CUSTNAME,CDES,IVDATE,QPRICE,VAT,TOTPRICE,DETAILS,CODEDES"
            f"&$orderby=IVDATE desc"
            f"&$top={top}"
        )
        resp = http_requests.get(api_url, headers=headers, auth=auth, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        invoices = []
        for row in data.get("value", []):
            invoices.append({
                "ivnum": row.get("IVNUM", ""),
                "customer": row.get("CUSTNAME", ""),
                "customerName": row.get("CDES", ""),
                "date": row.get("IVDATE", ""),
                "priceBeforeVat": row.get("QPRICE", 0),
                "vat": row.get("VAT", 0),
                "totalPrice": row.get("TOTPRICE", 0),
                "details": row.get("DETAILS", ""),
                "site": row.get("CODEDES", ""),
            })

        return jsonify({"ok": True, "invoices": invoices})
    except Exception as e:
        logger.error(f"Customer invoices fetch failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/cinvoice-download", methods=["POST"])
def download_cinvoice_pdf():
    """Download CINVOICE PDF attachment from Priority."""
    set_priority_env()
    data = request.get_json(silent=True) or {}
    ivnum = data.get("ivnum", "").strip()
    if not ivnum:
        return jsonify({"error": "Missing ivnum"}), 400

    try:
        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        headers = {"Accept": "application/json", "OData-Version": "4.0"}

        att_url = f"{url}/CINVOICES(IVNUM='{ivnum}',IVTYPE='C',DEBIT='D')/EXTFILES_SUBFORM"
        resp = http_requests.get(att_url, headers=headers, auth=auth, timeout=30)
        resp.raise_for_status()

        attachments = resp.json().get("value", [])
        if not attachments:
            return jsonify({"error": f"לא נמצא נספח לחשבונית {ivnum}"}), 404

        att = attachments[0]
        raw = att.get("EXTFILENAME", "")
        suffix = att.get("SUFFIX", "pdf")

        mime_type = "application/pdf"
        if raw.startswith("data:"):
            header, b64_data = raw.split(",", 1) if "," in raw else ("", raw)
            if ";" in header:
                mime_type = header.split(":")[1].split(";")[0]
        else:
            b64_data = raw

        import base64
        file_bytes = base64.b64decode(b64_data)
        safe_name = ivnum.replace("/", "-").replace("\\", "-")
        ext = suffix if suffix.startswith(".") else f".{suffix}"
        filename = f"{safe_name}{ext}"

        return send_file(
            io.BytesIO(file_bytes),
            mimetype=mime_type,
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        logger.error(f"CINVOICE download failed for {ivnum}: {e}")
        return jsonify({"error": f"שגיאה בהורדת חשבונית: {e}"}), 500


@app.route("/api/hr/send-invoices-email", methods=["POST"])
def send_invoices_email():
    """Download selected CINVOICE PDFs and send them as email attachments."""
    set_priority_env()
    data = request.get_json(force=True)
    email_to = data.get("email", "").strip()
    invoice_nums = data.get("invoices", [])
    customer_name = data.get("customerName", "")

    if not email_to:
        return jsonify({"ok": False, "error": "Missing email"}), 400
    if not invoice_nums:
        return jsonify({"ok": False, "error": "No invoices selected"}), 400

    try:
        import base64
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email.mime.text import MIMEText
        from email import encoders
        import boto3

        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        headers_api = {"Accept": "application/json", "OData-Version": "4.0"}

        # Build email
        msg = MIMEMultipart()
        msg["Subject"] = f"חשבוניות — {customer_name}" if customer_name else "חשבוניות"
        msg["From"] = os.getenv("GMAIL_USER", "arielmpinvoice@gmail.com")
        msg["To"] = email_to

        body_text = f"מצורפות {len(invoice_nums)} חשבוניות"
        if customer_name:
            body_text += f" של {customer_name}"
        msg.attach(MIMEText(body_text, "plain", "utf-8"))

        attached = 0
        errors = []
        for ivnum in invoice_nums:
            try:
                att_url = f"{url}/CINVOICES(IVNUM='{ivnum}',IVTYPE='C',DEBIT='D')/EXTFILES_SUBFORM"
                resp = http_requests.get(att_url, headers=headers_api, auth=auth, timeout=30)
                resp.raise_for_status()
                attachments = resp.json().get("value", [])
                if not attachments:
                    errors.append(f"{ivnum}: לא נמצא נספח")
                    continue

                att = attachments[0]
                raw = att.get("EXTFILENAME", "")
                suffix = att.get("SUFFIX", "pdf")
                if raw.startswith("data:"):
                    _, b64_data = raw.split(",", 1) if "," in raw else ("", raw)
                else:
                    b64_data = raw

                file_bytes = base64.b64decode(b64_data)
                ext = suffix if suffix.startswith(".") else f".{suffix}"
                safe_name = ivnum.replace("/", "-").replace("\\", "-")
                filename = f"{safe_name}{ext}"

                part = MIMEBase("application", "octet-stream")
                part.set_payload(file_bytes)
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={filename}")
                msg.attach(part)
                attached += 1
            except Exception as e:
                errors.append(f"{ivnum}: {e}")

        if attached == 0:
            return jsonify({"ok": False, "error": f"לא הצלחתי לצרף חשבוניות. {'; '.join(errors)}"}), 500

        # Send via Gmail SMTP
        import smtplib
        gmail_user = os.getenv("GMAIL_USER", "arielmpinvoice@gmail.com")
        gmail_pass = os.getenv("GMAIL_APP_PASSWORD", "")
        msg["From"] = gmail_user

        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.starttls()
            smtp.login(gmail_user, gmail_pass)
            smtp.send_message(msg)

        logger.info(f"Sent {attached} invoices to {email_to}")
        return jsonify({"ok": True, "sent": attached, "errors": errors})
    except Exception as e:
        logger.error(f"Send invoices email failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Delivery Notes Manager ───────────────────────────────────

@app.route("/api/hr/customer-delivery-notes", methods=["GET"])
def get_customer_delivery_notes():
    """Fetch last N DOCUMENTS_D for a customer (branch 102, finalized)."""
    try:
        customer = request.args.get("customer", "").strip()
        if customer.endswith("-102"):
            customer = customer[:-4]
        if not customer:
            return jsonify({"ok": False, "error": "Missing customer"}), 400

        top = int(request.args.get("top", "10"))
        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        headers = {"Accept": "application/json", "OData-Version": "4.0"}

        api_url = (
            f"{url}/DOCUMENTS_D"
            f"?$filter=BRANCHNAME eq '102' and CUSTNAME eq '{customer}' and STATDES eq 'סופית'"
            f"&$select=DOCNO,CUSTNAME,CDES,CURDATE,QPRICE,VAT,TOTPRICE,DETAILS,CODEDES,IVALL"
            f"&$orderby=CURDATE desc"
            f"&$top={top}"
        )
        resp = http_requests.get(api_url, headers=headers, auth=auth, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        docs = []
        for row in data.get("value", []):
            docs.append({
                "docno": row.get("DOCNO", ""),
                "customer": row.get("CUSTNAME", ""),
                "customerName": row.get("CDES", ""),
                "date": row.get("CURDATE", ""),
                "priceBeforeVat": row.get("QPRICE", 0),
                "vat": row.get("VAT", 0),
                "totalPrice": row.get("TOTPRICE", 0),
                "details": row.get("DETAILS", ""),
                "site": row.get("CODEDES", ""),
                "charged": row.get("IVALL") == "Y",
            })

        return jsonify({"ok": True, "docs": docs})
    except Exception as e:
        logger.error(f"Customer delivery notes fetch failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/delivery-note-download", methods=["POST"])
def download_delivery_note_pdf():
    """Download DOCUMENTS_D PDF attachment from Priority."""
    set_priority_env()
    data = request.get_json(silent=True) or {}
    docno = data.get("docno", "").strip()
    if not docno:
        return jsonify({"error": "Missing docno"}), 400

    try:
        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        headers_api = {"Accept": "application/json", "OData-Version": "4.0"}

        att_url = f"{url}/DOCUMENTS_D(DOCNO='{docno}',TYPE='D')/EXTFILES_SUBFORM"
        resp = http_requests.get(att_url, headers=headers_api, auth=auth, timeout=30)
        resp.raise_for_status()

        attachments = resp.json().get("value", [])
        if not attachments:
            return jsonify({"error": f"לא נמצא נספח לתעודה {docno}"}), 404

        att = attachments[0]
        raw = att.get("EXTFILENAME", "")
        suffix = att.get("SUFFIX", "pdf")

        import base64
        mime_type = "application/pdf"
        if raw.startswith("data:"):
            header, b64_data = raw.split(",", 1) if "," in raw else ("", raw)
            if ";" in header:
                mime_type = header.split(":")[1].split(";")[0]
        else:
            b64_data = raw

        file_bytes = base64.b64decode(b64_data)
        safe_name = docno.replace("/", "-").replace("\\", "-")
        ext = suffix if suffix.startswith(".") else f".{suffix}"
        filename = f"{safe_name}{ext}"

        return send_file(
            io.BytesIO(file_bytes),
            mimetype=mime_type,
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        logger.error(f"Delivery note download failed for {docno}: {e}")
        return jsonify({"error": f"שגיאה בהורדת תעודה: {e}"}), 500


@app.route("/api/hr/send-delivery-notes-email", methods=["POST"])
def send_delivery_notes_email():
    """Send selected delivery note PDFs via email."""
    set_priority_env()
    data = request.get_json(force=True)
    email_to = data.get("email", "").strip()
    doc_nums = data.get("docs", [])
    customer_name = data.get("customerName", "")

    if not email_to:
        return jsonify({"ok": False, "error": "Missing email"}), 400
    if not doc_nums:
        return jsonify({"ok": False, "error": "No docs selected"}), 400

    try:
        import base64
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email.mime.text import MIMEText
        from email import encoders
        import smtplib

        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        headers_api = {"Accept": "application/json", "OData-Version": "4.0"}

        gmail_user = os.getenv("GMAIL_USER", "arielmpinvoice@gmail.com")
        gmail_pass = os.getenv("GMAIL_APP_PASSWORD", "")

        msg = MIMEMultipart()
        msg["Subject"] = f"תעודות משלוח — {customer_name}" if customer_name else "תעודות משלוח"
        msg["From"] = gmail_user
        msg["To"] = email_to

        body_text = f"מצורפות {len(doc_nums)} תעודות משלוח"
        if customer_name:
            body_text += f" של {customer_name}"
        msg.attach(MIMEText(body_text, "plain", "utf-8"))

        attached = 0
        for docno in doc_nums:
            try:
                att_url = f"{url}/DOCUMENTS_D(DOCNO='{docno}',TYPE='D')/EXTFILES_SUBFORM"
                resp = http_requests.get(att_url, headers=headers_api, auth=auth, timeout=30)
                resp.raise_for_status()
                attachments = resp.json().get("value", [])
                if not attachments:
                    continue

                att = attachments[0]
                raw = att.get("EXTFILENAME", "")
                suffix = att.get("SUFFIX", "pdf")
                if raw.startswith("data:"):
                    _, b64_data = raw.split(",", 1) if "," in raw else ("", raw)
                else:
                    b64_data = raw

                file_bytes = base64.b64decode(b64_data)
                ext = suffix if suffix.startswith(".") else f".{suffix}"
                safe_name = docno.replace("/", "-").replace("\\", "-")
                filename = f"{safe_name}{ext}"

                part = MIMEBase("application", "octet-stream")
                part.set_payload(file_bytes)
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={filename}")
                msg.attach(part)
                attached += 1
            except Exception as e:
                logger.error(f"Attach {docno} failed: {e}")

        if attached == 0:
            return jsonify({"ok": False, "error": "לא הצלחתי לצרף תעודות"}), 500

        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.starttls()
            smtp.login(gmail_user, gmail_pass)
            smtp.send_message(msg)

        logger.info(f"Sent {attached} delivery notes to {email_to}")
        return jsonify({"ok": True, "sent": attached})
    except Exception as e:
        logger.error(f"Send delivery notes email failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Invoice Printer ──────────────────────────────────────────

@app.route("/api/invoice-printer/download", methods=["POST"])
def download_invoice_pdf():
    """Download invoice PDF attachment from Priority."""
    set_priority_env()
    data = request.get_json(silent=True) or {}
    ivnum = data.get("ivnum", "").strip()

    if not ivnum:
        return jsonify({"error": "Missing ivnum"}), 400

    try:
        file_bytes, filename, mime_type = invoice_printer.get_invoice_attachment_bytes(ivnum)
    except Exception as e:
        logger.error(f"Error fetching invoice {ivnum}: {e}")
        return jsonify({"error": f"שגיאה בשליפת חשבונית: {e}"}), 500

    if file_bytes is None:
        return jsonify({"error": f"לא נמצא נספח לחשבונית {ivnum}"}), 404

    # Save local copy when running locally (not on Lambda)
    if os.environ.get("IS_LAMBDA") != "true":
        try:
            local_dir = Path(r"C:\Users\User\Documents\חשבוניות פריורטי")
            local_dir.mkdir(parents=True, exist_ok=True)
            local_path = local_dir / filename
            with open(local_path, "wb") as f:
                f.write(file_bytes)
            logger.info(f"Saved local copy: {local_path}")
        except Exception as e:
            logger.warning(f"Could not save local copy: {e}")

    return send_file(
        io.BytesIO(file_bytes),
        mimetype=mime_type,
        as_attachment=True,
        download_name=filename,
    )


# ── Invoice Analyzer (Claude AI) ─────────────────────────────

@app.route("/api/analyze-invoice", methods=["POST"])
def analyze_invoice_pdf():
    """Analyze supplier invoice PDF using Claude Vision AI."""
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file uploaded"}), 400

    uploaded = request.files["file"]
    pdf_bytes = uploaded.read()

    if not pdf_bytes:
        return jsonify({"ok": False, "error": "Empty file"}), 400

    result = llm2000_analyzer.analyze_pdf(pdf_bytes)

    if not result.get("ok"):
        status = 500 if "API error" in result.get("error", "") else 400
        return jsonify(result), status

    # Enrich invoices with supplier data from cache
    cache = _load_supplier_cache()
    for inv in result.get("invoices", []):
        cid = inv.get("companyId", "")
        if cid and cid in cache:
            inv["supplier"] = cache[cid]["supname"]
            inv["supplierName"] = cache[cid]["supdes"]

    return jsonify(result)


# ── Supplier Cache ───────────────────────────────────────────
# Maps company ID (ח.פ.) → {supname, supdes} for auto-fill.
# Grows over time as users confirm supplier mappings.

_SUPPLIER_CACHE_PATH = PROJECT_ROOT / "backend" / "supplier_cache.json"


def _load_supplier_cache():
    """Load supplier cache from JSON file."""
    if _SUPPLIER_CACHE_PATH.exists():
        try:
            return json.loads(_SUPPLIER_CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_supplier_cache(cache):
    """Save supplier cache to JSON file."""
    _SUPPLIER_CACHE_PATH.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


@app.route("/api/supplier-mapping", methods=["POST"])
def save_supplier_mapping():
    """Save a company ID → supplier mapping for future auto-fill."""
    data = request.get_json()
    company_id = data.get("companyId", "").strip()
    supname = data.get("supname", "").strip()
    supdes = data.get("supdes", "").strip()

    if not company_id or not supname:
        return jsonify({"ok": False, "error": "companyId and supname required"}), 400

    cache = _load_supplier_cache()
    cache[company_id] = {"supname": supname, "supdes": supdes}
    _save_supplier_cache(cache)

    return jsonify({"ok": True, "saved": {company_id: cache[company_id]}})


@app.route("/api/suppliers", methods=["GET"])
def list_suppliers():
    """Fetch all unique suppliers from Priority YINVOICES for dropdown."""
    set_priority_env()
    url = get_priority_url()
    auth = HTTPBasicAuth(
        os.getenv("PRIORITY_USERNAME", ""),
        os.getenv("PRIORITY_PASSWORD", ""),
    )
    headers = {"Accept": "application/json", "OData-Version": "4.0"}

    suppliers = {}
    next_url = f"{url}/YINVOICES?$select=SUPNAME,CDES&$orderby=SUPNAME"
    pages = 0
    while next_url and pages < 30:
        resp = http_requests.get(next_url, headers=headers, auth=auth)
        if resp.status_code != 200:
            break
        data = resp.json()
        for row in data.get("value", []):
            sup = row.get("SUPNAME")
            if sup and sup not in suppliers:
                suppliers[sup] = row.get("CDES", "")
        next_url = data.get("@odata.nextLink")
        pages += 1

    result = [{"supname": k, "supdes": v} for k, v in sorted(suppliers.items())]
    return jsonify({"ok": True, "suppliers": result, "count": len(result)})


# ── Bot Scripts ──────────────────────────────────────────────

@app.route("/api/bot-scripts/generate", methods=["POST"])
def generate_bot_script():
    """Use Claude AI to generate a bot script from a plain-language description."""
    import re
    data = request.get_json(silent=True) or {}
    description = data.get("description", "").strip()
    if not description:
        return jsonify({"ok": False, "error": "description required"}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY not configured"}), 500

    model = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")

    prompt = f"""You are building a WhatsApp bot conversation script for a Hebrew-speaking maintenance company.
Generate a bot script in JSON format based on this description:

{description}

The JSON must follow this EXACT schema:
{{
  "name": "Script name in Hebrew",
  "greeting_known": "Opening message, may use {{{{customer_name}}}} placeholder",
  "greeting_unknown": "Opening message for unknown customers",
  "first_step": "STEP_1",
  "steps": [
    {{
      "id": "STEP_1",
      "type": "text_input",
      "text": "Question text in Hebrew",
      "save_to": "field_name_in_english",
      "next_step": "STEP_2"
    }},
    {{
      "id": "STEP_2",
      "type": "buttons",
      "text": "Question text in Hebrew",
      "buttons": [
        {{"id": "btn_2_1", "title": "Button text (max 20 chars)", "next_step": "DONE_1"}},
        {{"id": "btn_2_2", "title": "Button text (max 20 chars)", "next_step": "DONE_1"}}
      ]
    }}
  ],
  "done_actions": {{
    "DONE_1": {{
      "text": "Closing message in Hebrew",
      "action": "save_service_call"
    }}
  }}
}}

Rules:
- All questions and messages must be in Hebrew
- Button titles: max 20 characters, max 3 buttons per step
- save_to field names in English (e.g. description, location, category, urgency, phone, name)
- action must be: save_service_call OR save_message
- Steps connect linearly: STEP_1 → STEP_2 → ... → DONE_1
- Return ONLY the raw JSON object, no markdown, no explanation"""

    try:
        resp = http_requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": model,
                "max_tokens": 2000,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=30,
        )
        if resp.status_code != 200:
            return jsonify({"ok": False, "error": f"AI error: {resp.status_code}"}), 500

        text = resp.json()["content"][0]["text"].strip()
        # Strip markdown code fences if present
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        script_data = json.loads(text)
        return jsonify({"ok": True, "script": script_data})
    except json.JSONDecodeError as e:
        return jsonify({"ok": False, "error": f"AI returned invalid JSON: {e}"}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot-sessions", methods=["GET"])
def api_bot_sessions():
    """List recent bot sessions with their activity logs (for diagnostics)."""
    try:
        sessions = troubleshoot_sessions_db.list_sessions(limit=50)
        sessions.sort(key=lambda s: s.get("created_at", ""), reverse=True)
        light = []
        for s in sessions:
            light.append({
                "phone": s.get("phone"),
                "name": s.get("name"),
                "customer_name": s.get("customer_name"),
                "device_number": s.get("device_number"),
                "script_id": s.get("script_id"),
                "step": s.get("step"),
                "status": s.get("status", "active"),
                "created_at": s.get("created_at"),
                "updated_at": s.get("updated_at"),
                "session_log": s.get("session_log", []),
            })
        return jsonify({"ok": True, "sessions": light})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot-scripts", methods=["GET"])
def list_bot_scripts():
    """List all bot conversation scripts."""
    try:
        scripts = bot_scripts_db.list_scripts()
        return jsonify({"ok": True, "scripts": scripts, "count": len(scripts)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot-scripts/<script_id>", methods=["GET"])
def get_bot_script(script_id):
    """Get a single bot script by ID."""
    try:
        script = bot_scripts_db.get_script(script_id, use_cache=False)
        if not script:
            return jsonify({"ok": False, "error": "תסריט לא נמצא"}), 404
        return jsonify({"ok": True, "script": script})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot-scripts", methods=["POST"])
def create_bot_script():
    """Create a new bot script."""
    data = request.get_json(silent=True) or {}
    if not data.get("script_id") or not data.get("name"):
        return jsonify({"ok": False, "error": "script_id and name are required"}), 400
    try:
        result = bot_scripts_db.save_script(data)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot-scripts/<script_id>", methods=["PUT"])
def update_bot_script(script_id):
    """Update an existing bot script."""
    data = request.get_json(silent=True) or {}
    data["script_id"] = script_id
    try:
        result = bot_scripts_db.save_script(data)
        # Invalidate the M10010 engine cache too
        bot_scripts_db.invalidate_cache(script_id)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Bot Prompts (LLM Training) ───────────────────────────────

@app.route("/api/bot-prompts", methods=["GET"])
def list_bot_prompts():
    """List all LLM prompts."""
    try:
        prompts = bot_prompts_db.list_prompts()
        return jsonify({"ok": True, "prompts": prompts, "count": len(prompts)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot-prompts/active", methods=["GET"])
def get_active_bot_prompt():
    """Get the currently active LLM prompt."""
    try:
        prompt = bot_prompts_db.get_active_prompt(use_cache=False)
        if not prompt:
            return jsonify({"ok": False, "error": "No active prompt found"}), 404
        return jsonify({"ok": True, "prompt": prompt})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot-prompts", methods=["POST"])
def create_bot_prompt():
    """Create a new LLM prompt."""
    data = request.get_json(silent=True) or {}
    if not data.get("prompt_id") or not data.get("content"):
        return jsonify({"ok": False, "error": "prompt_id and content are required"}), 400
    try:
        result = bot_prompts_db.save_prompt(data)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bot-prompts/<prompt_id>", methods=["PUT"])
def update_bot_prompt(prompt_id):
    """Update an existing LLM prompt."""
    data = request.get_json(silent=True) or {}
    data["prompt_id"] = prompt_id
    try:
        result = bot_prompts_db.save_prompt(data)
        bot_prompts_db.invalidate_cache(prompt_id)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Conversation History ─────────────────────────────────────

@app.route("/api/conversations", methods=["GET"])
def get_conversations():
    """Get conversation history (service calls)."""
    try:
        status = request.args.get("status")
        phone = request.args.get("phone")
        limit = int(request.args.get("limit", 50))
        calls = maint_db.get_service_calls(status=status, phone=phone, limit=limit)
        return jsonify({"ok": True, "conversations": calls, "count": len(calls)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/conversations/<call_id>", methods=["GET"])
def get_conversation_detail(call_id):
    """Get full detail of a single conversation/service call."""
    try:
        call = maint_db.get_service_call(call_id)
        if not call:
            return jsonify({"ok": False, "error": "Conversation not found"}), 404
        return jsonify({"ok": True, "conversation": call})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Knowledge Base (RAG) ─────────────────────────────────────

@app.route("/api/knowledge", methods=["GET"])
def list_knowledge():
    """List all active knowledge items."""
    try:
        item_type = request.args.get("type")
        items = knowledge_db.list_items(item_type=item_type)
        return jsonify({"ok": True, "items": items, "count": len(items)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/knowledge/<item_id>", methods=["GET"])
def get_knowledge_item(item_id):
    """Get a single knowledge item."""
    try:
        item = knowledge_db.get_item(item_id)
        if not item:
            return jsonify({"ok": False, "error": "פריט ידע לא נמצא"}), 404
        # Strip embedding from response
        item.pop("embedding", None)
        return jsonify({"ok": True, "item": item})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/knowledge", methods=["POST"])
def create_knowledge_item():
    """Create a new knowledge item with auto-embedding."""
    data = request.get_json(silent=True) or {}
    if not data.get("title") or not data.get("content"):
        return jsonify({"ok": False, "error": "title and content are required"}), 400
    try:
        # Generate embedding
        embedding = rag_retrieval.generate_embedding(data["content"])
        if embedding:
            data["embedding"] = embedding
        else:
            logger.warning("Failed to generate embedding, saving without it")

        result = knowledge_db.save_item(data)
        return jsonify({"ok": True, **result, "has_embedding": embedding is not None})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/knowledge/<item_id>", methods=["PUT"])
def update_knowledge_item(item_id):
    """Update a knowledge item (re-embeds if content changed)."""
    data = request.get_json(silent=True) or {}
    data["id"] = item_id
    try:
        # Re-generate embedding if content provided
        if data.get("content"):
            embedding = rag_retrieval.generate_embedding(data["content"])
            if embedding:
                data["embedding"] = embedding

        result = knowledge_db.save_item(data)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/knowledge/<item_id>", methods=["DELETE"])
def delete_knowledge_item(item_id):
    """Deactivate a knowledge item."""
    try:
        knowledge_db.delete_item(item_id)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/knowledge/from-call", methods=["POST"])
def create_knowledge_from_call():
    """Create a knowledge item from an existing service call conversation."""
    data = request.get_json(silent=True) or {}
    call_id = data.get("call_id")
    notes = data.get("notes", "")
    if not call_id:
        return jsonify({"ok": False, "error": "call_id is required"}), 400
    try:
        # Load the service call
        call = maint_db.get_service_call(call_id)
        if not call:
            return jsonify({"ok": False, "error": "קריאת שירות לא נמצאה"}), 404

        # Build knowledge content from call data + operator notes
        parts = []
        if call.get("issue_type"):
            parts.append(f"סוג תקלה: {call['issue_type']}")
        if call.get("description"):
            parts.append(f"תיאור: {call['description']}")
        if call.get("summary"):
            parts.append(f"תמצית: {call['summary']}")
        if call.get("location"):
            parts.append(f"מיקום: {call['location']}")
        if notes:
            parts.append(f"הערות מפעיל: {notes}")

        content = "\n".join(parts)
        title = f"{call.get('issue_type', 'שיחה')} - {call.get('cdes', call.get('name', 'לקוח'))}"

        item_data = {
            "type": "feedback",
            "title": title,
            "content": content,
            "source_call_id": call_id,
            "tags": [call.get("urgency", ""), call.get("branchname", "")],
        }

        # Generate embedding
        embedding = rag_retrieval.generate_embedding(content)
        if embedding:
            item_data["embedding"] = embedding

        result = knowledge_db.save_item(item_data)
        return jsonify({"ok": True, **result, "title": title})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Health ───────────────────────────────────────────────────

@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"ok": True, "service": "urbangroup-backend"})


# ── SharePoint / HR Management ──────────────────────────────

_sp_connector = None

def _get_sp_connector():
    global _sp_connector
    if _sp_connector is None:
        sp_path = PROJECT_ROOT / "agents" / "tools-connection" / "5100-sharepoint" / "5100-sharepoint_connector.py"
        spec = importlib.util.spec_from_file_location("sp_connector_5100", sp_path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _sp_connector = mod
    return _sp_connector


HR_SHARE_URL = "https://yaelisrael.sharepoint.com/:x:/g/Realestateproject/IQCk9K_jvlY-S6n0U3Wls4rTAQDvNEL8m9GIHU16NXW-37E"

# ── Local HR persistence ──────────────────────────────────────────────
HR_LOCAL_DIR = Path("/tmp/hr_local") if IS_LAMBDA else \
    PROJECT_ROOT / "output" / "hr_local"


def _hr_local_path(sheet):
    """Return path to local JSON file for a given sheet."""
    safe = sheet.replace("/", "_").replace("\\", "_")
    return HR_LOCAL_DIR / f"{safe}.json"


def _read_local_hr(sheet):
    """Read local HR data for a sheet. Returns dict or None."""
    p = _hr_local_path(sheet)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None
    return None


def _write_local_hr(sheet, data):
    """Write local HR data for a sheet."""
    HR_LOCAL_DIR.mkdir(parents=True, exist_ok=True)
    p = _hr_local_path(sheet)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def _clear_local_hr(sheet):
    """Remove local HR data for a sheet (after successful SharePoint save)."""
    p = _hr_local_path(sheet)
    if p.exists():
        p.unlink()


# Known header markers for the main table (column C and D)
_HR_HEADER_MARKERS = {"לקוח", "אתר"}


def _find_main_table(all_rows):
    """Find header row index and last data row index in the sheet.

    Args:
        all_rows: 2D list of all cell values (from row 1 onward)

    Returns:
        (header_idx, last_data_idx) — 0-based indices into all_rows,
        or (None, None) if not found.
    """
    header_idx = None
    for i, row in enumerate(all_rows):
        # Header row has "לקוח" in col D (idx 3) and "אתר" in col F (idx 5)
        c3 = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        c5 = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        if c3 in _HR_HEADER_MARKERS and c5 in _HR_HEADER_MARKERS:
            header_idx = i
            break

    if header_idx is None:
        return None, None

    # Find last data row: scan from header+1 until we hit several consecutive empty rows
    last_data_idx = header_idx
    empty_streak = 0
    for i in range(header_idx + 1, len(all_rows)):
        row = all_rows[i]
        has_data = any(cell for cell in row[:24] if cell is not None and str(cell).strip())
        if has_data:
            last_data_idx = i
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak >= 5:
                break

    return header_idx, last_data_idx


@app.route("/api/hr/local-save", methods=["POST"])
def hr_local_save():
    """Auto-save HR table state locally so edits survive server/network issues.

    Body JSON:
        sheet: sheet name
        rows: full edited rows array
        dirtyKeys: list of "excelRow:colIdx" strings
        deletedRows: list of ROW_INDEX values
    """
    body = request.get_json(force=True)
    sheet = body.get("sheet", "2.26")
    try:
        _write_local_hr(sheet, {
            "rows": body.get("rows", []),
            "dirtyKeys": body.get("dirtyKeys", []),
            "deletedRows": body.get("deletedRows", []),
            "savedAt": _now_il().strftime("%Y-%m-%d %H:%M:%S"),
        })
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"HR local save failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/local-data", methods=["GET"])
def hr_local_data():
    """Return locally saved HR data for a sheet, if any."""
    sheet = request.args.get("sheet", "2.26")
    local = _read_local_hr(sheet)
    if local:
        return jsonify({"ok": True, "hasLocal": True, **local})
    return jsonify({"ok": True, "hasLocal": False})


def _parse_hr_sheet(all_rows):
    """Parse raw Excel rows into data_rows + filters. Returns (data_rows, filters) or (None, error_msg)."""
    header_idx, last_data_idx = _find_main_table(all_rows)
    if header_idx is None:
        return None, "לא נמצאה טבלה ראשית בגיליון"

    data_rows = []
    for i in range(header_idx + 1, last_data_idx + 1):
        row = all_rows[i]
        excel_row = i + 1
        if row[3] or row[5]:
            row.append(excel_row)
            data_rows.append(row)

    customers = sorted(set(str(r[3]).strip() for r in data_rows if r[3]))
    sites = sorted(set(str(r[5]).strip() for r in data_rows if r[5]))
    contractors = sorted(set(str(r[11]).strip() for r in data_rows if r[11]))

    customer_sites = {}
    for r in data_rows:
        cust = str(r[3]).strip() if r[3] else ""
        site = str(r[5]).strip() if r[5] else ""
        if cust and site:
            customer_sites.setdefault(cust, set()).add(site)
    customer_sites = {k: sorted(v) for k, v in customer_sites.items()}

    filters = {
        "customers": customers,
        "sites": sites,
        "contractors": contractors,
        "customer_sites": customer_sites,
    }
    return data_rows, filters


@app.route("/api/hr/db-data", methods=["GET"])
def get_hr_db_data():
    """Load HR sheet data from DynamoDB cache (fast startup)."""
    sheet = request.args.get("sheet", "2.26")
    try:
        cached = delivery_notes_db.load_hr_sheet(sheet)
        if not cached:
            return jsonify({"ok": False, "error": "no_cache"})
        return jsonify({
            "ok": True,
            "rows": cached["rows"],
            "filters": cached["filters"],
            "fromDb": True,
            "syncedAt": cached.get("synced_at", ""),
        })
    except Exception as e:
        logger.error(f"HR DB read failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/db-data", methods=["POST"])
def save_hr_db_data():
    """Save HR sheet data to DynamoDB cache (called by frontend after save)."""
    try:
        data = request.get_json(force=True)
        sheet = data.get("sheet", "2.26")
        rows = data.get("rows", [])
        filters = data.get("filters", {})
        delivery_notes_db.save_hr_sheet(sheet, rows, filters)
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"HR DB save failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/sheet-data", methods=["GET"])
def get_hr_sheet_data():
    """Read main table from the HR Excel file on SharePoint and save to DB cache."""
    sheet = request.args.get("sheet", "2.26")
    try:
        sp = _get_sp_connector()
        excel = sp.SharePointExcel(HR_SHARE_URL)

        data = excel.read(sheet, "A1:X1000")
        all_rows = data["values"]
        if not all_rows:
            return jsonify({"ok": True, "headers": [], "rows": [], "filters": {}})

        data_rows, filters = _parse_hr_sheet(all_rows)
        if data_rows is None:
            return jsonify({"ok": False, "error": filters})

        # Save to DB cache for fast startup
        try:
            delivery_notes_db.save_hr_sheet(sheet, data_rows, filters)
        except Exception as cache_err:
            logger.error(f"HR DB cache save failed (non-fatal): {cache_err}")

        return jsonify({
            "ok": True,
            "rows": data_rows,
            "filters": filters,
        })
    except Exception as e:
        logger.error(f"HR sheet read failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/sheets", methods=["GET"])
def get_hr_sheets():
    """List available worksheets - merged from SharePoint Excel and DB-only months."""
    sheets = []
    try:
        sp = _get_sp_connector()
        excel = sp.SharePointExcel(HR_SHARE_URL)
        sheets = excel.sheets()
    except Exception as e:
        logger.warning(f"SharePoint sheets fetch failed: {e}")
    # Add DB-only months (created via "Create Month" button)
    try:
        resp = delivery_notes_db._table.scan(
            FilterExpression="begins_with(id, :p)",
            ExpressionAttributeValues={":p": "HR_SHEET_"},
            ProjectionExpression="id",
        )
        for item in resp.get("Items", []):
            month = item.get("id", "").replace("HR_SHEET_", "")
            if month and month not in sheets:
                sheets.append(month)
    except Exception as e:
        logger.warning(f"DB sheets fetch failed: {e}")
    return jsonify({"ok": True, "sheets": sheets})


@app.route("/api/hr/save-changes", methods=["POST"])
def save_hr_changes():
    """Save changed cells back to the HR Excel file on SharePoint.

    Body JSON:
        sheet: sheet name (e.g. '2.26')
        changes: list of {row: int, col: int, value: any}
            row = Excel row number (38+), col = 0-based column index (A=0)
        newRows: list of [col0, col1, ..., col21] — new rows to append
    """
    body = request.get_json(force=True)
    sheet = body.get("sheet", "2.26")
    changes = body.get("changes", [])
    all_ordered_rows = body.get("allOrderedRows")

    if not changes and not all_ordered_rows:
        return jsonify({"ok": True, "updated": 0})
    logger.info(f"HR save: {len(changes)} changes, allOrderedRows={'yes' if all_ordered_rows else 'no'} ({len(all_ordered_rows) if all_ordered_rows else 0} rows)")

    try:
        sp = _get_sp_connector()
        excel = sp.SharePointExcel(HR_SHARE_URL)

        # Retry helper for SharePoint 409 conflicts
        import time
        def _sp_call(fn, retries=3, delay=2):
            for attempt in range(retries):
                try:
                    return fn()
                except Exception as e:
                    if ('409' in str(e) or 'Conflict' in str(e)) and attempt < retries - 1:
                        logger.warning(f"SharePoint conflict, retry {attempt + 1}/{retries}")
                        time.sleep(delay * (attempt + 1))
                    else:
                        raise

        updated = 0

        # MODE 1: Full rewrite — when rows were added/deleted, frontend sends all rows in order
        if all_ordered_rows:
            # Read Excel to find header position
            table_data = excel.read(sheet, "A1:X1000")
            all_excel = table_data.get("values", [])
            header_idx, last_data_idx = _find_main_table(all_excel)

            # Keep header rows (everything up to and including header)
            header_rows = all_excel[:header_idx + 1] if header_idx is not None else []

            # Apply cell-level changes to the ordered rows
            changes_map = {}
            for ch in changes:
                key = f"{ch['row']}:{ch['col']}"
                changes_map.setdefault(ch['row'], {})[ch['col']] = ch['value']

            # Build new data rows from frontend order, applying any pending changes
            data_rows = []
            for row in all_ordered_rows:
                padded = list(row)[:24] + [''] * max(0, 24 - len(row))
                data_rows.append(padded)

            # Combine: headers + data rows
            final = header_rows + data_rows
            write_range = f"A1:X{len(final)}"
            _sp_call(lambda: sp.write_excel_range(
                excel.drive_id, excel.item_id,
                sheet, write_range, final
            ))

            # Clear any leftover rows in Excel (if old table was longer)
            old_len = len(all_excel)
            new_len = len(final)
            if old_len > new_len:
                empty_rows = [[''] * 24] * (old_len - new_len)
                clear_range = f"A{new_len + 1}:X{old_len}"
                _sp_call(lambda: sp.write_excel_range(
                    excel.drive_id, excel.item_id,
                    sheet, clear_range, empty_rows
                ))

            updated = len(data_rows)
            return jsonify({"ok": True, "updated": updated})

        # MODE 2: Cell-level updates only (no structural changes)
        if changes:
            rows_map = {}
            for ch in changes:
                row_num = ch["row"]
                col_idx = ch["col"]
                value = ch["value"]
                if row_num not in rows_map:
                    rows_map[row_num] = {}
                rows_map[row_num][col_idx] = value

            min_row = min(rows_map.keys())
            max_row = max(rows_map.keys())

            read_range = f"A{min_row}:X{max_row}"
            try:
                current = excel.read(sheet, read_range)
                all_vals = current.get("values", [])
            except Exception:
                all_vals = []

            needed = max_row - min_row + 1
            while len(all_vals) < needed:
                all_vals.append([''] * 24)

            for row_num, cols in rows_map.items():
                row_idx = row_num - min_row
                row_data = list(all_vals[row_idx])
                while len(row_data) < 24:
                    row_data.append('')
                for col_idx, value in cols.items():
                    if col_idx < 24:
                        row_data[col_idx] = value
                all_vals[row_idx] = row_data
                updated += len(cols)

            _sp_call(lambda: sp.write_excel_range(
                excel.drive_id, excel.item_id,
                sheet, read_range, all_vals
            ))

        # Clear local pending data after successful SharePoint save
        _clear_local_hr(sheet)

        return jsonify({"ok": True, "updated": updated})
    except Exception as e:
        logger.error(f"HR save failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _fetch_customers_from_priority():
    """Fetch customers list from Priority ERP (real env) — filtered to branch 102."""
    url = PRIORITY_URL_REAL
    auth = HTTPBasicAuth(
        os.getenv("PRIORITY_USERNAME", ""),
        os.getenv("PRIORITY_PASSWORD", ""),
    )
    headers = {"Accept": "application/json", "OData-Version": "4.0"}

    customers = []
    skip = 0
    while True:
        api_url = f"{url}/ACCOUNTS_RECEIVABLE?$select=ACCNAME,ACCDES&$orderby=ACCNAME&$top=500&$skip={skip}"
        resp = http_requests.get(api_url, headers=headers, auth=auth, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        rows = data.get("value", [])
        if not rows:
            break
        for row in rows:
            acc = row.get("ACCNAME", "")
            if acc.endswith("-102"):
                customers.append({
                    "code": acc,
                    "name": row.get("ACCDES", ""),
                })
        skip += len(rows)
        if len(rows) < 500:
            break
    return customers


@app.route("/api/hr/customers", methods=["GET"])
def get_hr_customers():
    """Return customers from DB cache. If no cache, fetch from Priority."""
    try:
        cached = delivery_notes_db.load_customers_cache()
        if cached:
            return jsonify({"ok": True, "customers": cached["customers"], "syncedAt": cached.get("synced_at", "")})
        # No cache — fetch live and save
        customers = _fetch_customers_from_priority()
        delivery_notes_db.save_customers_cache(customers)
        return jsonify({"ok": True, "customers": customers})
    except Exception as e:
        logger.error(f"HR customers fetch failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _fetch_sites_from_priority():
    """Fetch all sites 100-2000 from Priority CUSTDESTS_ONE."""
    url = PRIORITY_URL_REAL
    auth = HTTPBasicAuth(
        os.getenv("PRIORITY_USERNAME", ""),
        os.getenv("PRIORITY_PASSWORD", ""),
    )
    headers = {"Accept": "application/json", "OData-Version": "4.0"}

    sites = []
    skip = 0
    while True:
        api_url = (
            f"{url}/CUSTDESTS_ONE"
            f"?$select=CODE,CODEDES,CUSTNAMEA,CUSTNAME,STATE"
            f"&$orderby=CODE&$top=500&$skip={skip}"
        )
        resp = http_requests.get(api_url, headers=headers, auth=auth, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        rows = data.get("value", [])
        if not rows:
            break
        for row in rows:
            code = row.get("CODE", "")
            try:
                code_num = int(code)
            except (ValueError, TypeError):
                continue
            if code_num < 100 or code_num > 2000:
                continue
            sites.append({
                "code": code,
                "name": row.get("CODEDES", ""),
                "custCode": row.get("CUSTNAMEA", ""),
                "custName": row.get("CUSTNAME", ""),
                "city": row.get("STATE", ""),
            })
        skip += len(rows)
        if len(rows) < 500:
            break
    return sites


@app.route("/api/hr/sites", methods=["GET"])
def get_hr_sites():
    """Return sites from DB cache, filtered by customer. If no cache, fetch from Priority."""
    try:
        customer = request.args.get("customer", "").strip()
        if customer.endswith("-102"):
            customer = customer[:-4]

        cached = delivery_notes_db.load_sites_cache()
        if not cached:
            # No cache — fetch and save
            all_sites = _fetch_sites_from_priority()
            delivery_notes_db.save_sites_cache(all_sites)
            cached = {"sites": all_sites}

        sites = cached["sites"]
        if customer:
            sites = [s for s in sites if s.get("custCode") == customer]

        return jsonify({"ok": True, "sites": sites})
    except Exception as e:
        logger.error(f"HR sites fetch failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


_is_lambda = os.environ.get("IS_LAMBDA") == "true"


def _fetch_parts_from_priority():
    """Fetch parts 100-199 from Priority via OData API."""
    url = PRIORITY_URL_REAL
    auth = HTTPBasicAuth(
        os.getenv("PRIORITY_USERNAME", ""),
        os.getenv("PRIORITY_PASSWORD", ""),
    )
    headers = {"Accept": "application/json", "OData-Version": "4.0"}

    parts = []
    skip = 0
    while True:
        api_url = (
            f"{url}/LOGPART"
            f"?$filter=PARTNAME ge '100' and PARTNAME lt '200'"
            f"&$select=PARTNAME,PARTDES,UNITNAME,SPEC20"
            f"&$orderby=PARTNAME&$top=500&$skip={skip}"
        )
        resp = http_requests.get(api_url, headers=headers, auth=auth, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        rows = data.get("value", [])
        if not rows:
            break
        for row in rows:
            parts.append({
                "code": row.get("PARTNAME", ""),
                "name": row.get("PARTDES", ""),
                "unit": row.get("UNITNAME", ""),
                "spec20": row.get("SPEC20", ""),
            })
        skip += len(rows)
        if len(rows) < 500:
            break

    logger.info(f"Fetched {len(parts)} parts from Priority")
    return parts


def _save_parts_cache(parts):
    """Save parts to DynamoDB cache."""
    delivery_notes_db.save_parts_cache(parts)


def _load_parts_cache():
    """Load parts from DynamoDB cache."""
    data = delivery_notes_db.load_parts_cache()
    if data:
        return {"parts": data["parts"], "syncedAt": data.get("synced_at", "")}
    return None


@app.route("/api/hr/parts", methods=["GET"])
def get_hr_parts():
    """Return cached parts list. If no cache, try to fetch from Priority."""
    try:
        cache = _load_parts_cache()
        if cache:
            return jsonify({"ok": True, "parts": cache["parts"], "syncedAt": cache.get("syncedAt", "")})
        # No cache — try fetch live
        try:
            parts = _fetch_parts_from_priority()
            _save_parts_cache(parts)
            return jsonify({"ok": True, "parts": parts})
        except Exception as ex:
            logger.error(f"Parts live fetch failed: {ex}")
            return jsonify({"ok": False, "error": f"שגיאה בטעינת מקטים: {ex}"}), 200
    except Exception as e:
        logger.error(f"HR parts fetch failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/sync-priority", methods=["GET"])
def sync_hr_priority():
    """Fetch customers and suppliers from Priority ERP (real env, branch 102)."""
    try:
        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        headers = {"Accept": "application/json", "OData-Version": "4.0"}

        # 1. Fetch customers (use $skip — Priority doesn't return nextLink)
        customers = []
        skip = 0
        while True:
            api_url = f"{url}/CUSTOMERS?$select=CUSTNAME,CUSTDES&$orderby=CUSTNAME&$top=500&$skip={skip}"
            resp = http_requests.get(api_url, headers=headers, auth=auth, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            rows = data.get("value", [])
            if not rows:
                break
            for row in rows:
                customers.append({
                    "code": row.get("CUSTNAME", ""),
                    "name": row.get("CUSTDES", ""),
                })
            skip += len(rows)
            if len(rows) < 500:
                break

        # 2. Fetch suppliers (from YINVOICES — unique SUPNAME + CDES pairs)
        sup_map = {}
        next_url = f"{url}/YINVOICES?$select=SUPNAME,CDES&$top=500"
        while next_url:
            resp = http_requests.get(next_url, headers=headers, auth=auth, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            for row in data.get("value", []):
                code = row.get("SUPNAME", "")
                if code and code not in sup_map:
                    sup_map[code] = row.get("CDES", "")
            next_url = data.get("@odata.nextLink")
        suppliers = [{"code": k, "name": v} for k, v in sorted(sup_map.items())]

        # 3. Fetch parts 100-199 and cache
        parts_error = None
        try:
            parts = _fetch_parts_from_priority()
            _save_parts_cache(parts)
        except Exception as pe:
            logger.error(f"Parts sync failed (non-fatal): {pe}")
            parts_error = str(pe)

        # 4. Fetch HR customers (ACCOUNTS_RECEIVABLE -102) and cache
        hr_customers_error = None
        try:
            hr_customers = _fetch_customers_from_priority()
            delivery_notes_db.save_customers_cache(hr_customers)
        except Exception as ce:
            logger.error(f"HR customers sync failed (non-fatal): {ce}")
            hr_customers_error = str(ce)

        # 5. Fetch sites (CUSTDESTS_ONE 100-2000) and cache
        sites_error = None
        sites_count = 0
        try:
            all_sites = _fetch_sites_from_priority()
            delivery_notes_db.save_sites_cache(all_sites)
            sites_count = len(all_sites)
        except Exception as se:
            logger.error(f"Sites sync failed (non-fatal): {se}")
            sites_error = str(se)

        return jsonify({
            "ok": True,
            "customers": customers,
            "suppliers": suppliers,
            "partsCount": len(parts) if not parts_error else 0,
            "partsError": parts_error,
            "hrCustomersCount": len(hr_customers) if not hr_customers_error else 0,
            "sitesCount": sites_count,
            "syncedAt": _now_il().strftime("%Y-%m-%d %H:%M:%S"),
        })
    except Exception as e:
        logger.error(f"HR Priority sync failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/tasks", methods=["GET"])
def list_tasks():
    """List tasks, optionally filtered by status."""
    try:
        status = request.args.get("status")
        tasks = delivery_notes_db.list_tasks(status=status)
        return jsonify({"ok": True, "tasks": tasks})
    except Exception as e:
        logger.error(f"List tasks failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/tasks", methods=["POST"])
def create_task():
    """Create a new task."""
    try:
        data = request.get_json(force=True)
        description = data.get("description", "").strip()
        if not description:
            return jsonify({"ok": False, "error": "Missing description"}), 400
        month = data.get("month", "").strip()
        result = delivery_notes_db.save_task(description, month=month)
        return jsonify({"ok": True, "id": result["id"]})
    except Exception as e:
        logger.error(f"Create task failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/tasks/<task_id>", methods=["PUT"])
def update_task(task_id):
    """Update a task (toggle status, edit description)."""
    try:
        data = request.get_json(force=True)
        updates = {}
        for field in ("status", "description", "month"):
            if field in data:
                updates[field] = data[field]
        if updates:
            delivery_notes_db.update_task(task_id, updates)
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"Update task failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/tasks/<task_id>", methods=["DELETE"])
def delete_task(task_id):
    """Delete a task."""
    try:
        delivery_notes_db.delete_task(task_id)
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"Delete task failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/create-month", methods=["POST"])
def create_hr_month():
    """Create a new HR month, optionally copying data from a source month.

    Body:
      newMonth: e.g. "4.26"
      sourceMonth: optional, e.g. "3.26" - month to copy from
    """
    try:
        body = request.get_json(force=True)
        new_month = (body.get("newMonth") or "").strip()
        source_month = (body.get("sourceMonth") or "").strip()

        if not new_month:
            return jsonify({"ok": False, "error": "Missing newMonth"}), 400

        # Check if new month already exists
        existing = delivery_notes_db.load_hr_sheet(new_month)
        if existing:
            return jsonify({"ok": False, "error": f"Month {new_month} already exists"}), 400

        rows = []
        filters = {}
        # If source specified, copy from source
        if source_month:
            source_data = delivery_notes_db.load_hr_sheet(source_month)
            if not source_data:
                return jsonify({"ok": False, "error": f"Source month {source_month} not found"}), 400

            # Column indices to clear in copied rows
            CLEAR_COLS = {
                1,    # TRACKING - מעקב
                4,    # FILLING - מילוי
                12,   # HOURS_REG - שעות רגילות
                13,   # HOURS_125 - שעות 125%
                14,   # HOURS_150 - שעות 150%
                18,   # CUST_TOTAL - calculated, will be recalced
                22,   # CONT_TOTAL - calculated
                23,   # GAP - calculated
            }

            for row in source_data.get("rows", []):
                new_row = list(row)
                # Pad to 25 columns to be safe
                while len(new_row) < 25:
                    new_row.append('')
                for c in CLEAR_COLS:
                    if c < len(new_row):
                        new_row[c] = ''
                rows.append(new_row)
            filters = source_data.get("filters", {})

        # Save new month sheet
        delivery_notes_db.save_hr_sheet(new_month, rows, filters)

        # Copy contractor payments and clear finalAmount
        if source_month:
            cp = delivery_notes_db.get_contractor_payments(source_month)
            if cp and cp.get("data"):
                data = cp["data"]
                if isinstance(data, str):
                    import json as _json
                    data = _json.loads(data)
                new_data = []
                for p in data:
                    np = dict(p)
                    np["finalAmount"] = ""
                    np["afterTaxDeduction"] = ""
                    np["withVat"] = ""
                    np["payToday"] = ""
                    np["finalGreen"] = False
                    np["payTodayGreen"] = False
                    new_data.append(np)
                delivery_notes_db.save_contractor_payments(new_month, new_data)

        return jsonify({"ok": True, "newMonth": new_month, "rowCount": len(rows)})
    except Exception as e:
        logger.error(f"Create HR month failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/contractor-payments", methods=["GET"])
def get_contractor_payments():
    """Get saved contractor payments for a sheet."""
    sheet = request.args.get("sheet", "")
    if not sheet:
        return jsonify({"ok": False, "error": "Missing sheet"}), 400
    try:
        result = delivery_notes_db.get_contractor_payments(sheet)
        if result and result.get("data"):
            data = result["data"]
            if isinstance(data, str):
                import json as _json
                data = _json.loads(data)
            return jsonify({"ok": True, "data": data})
        return jsonify({"ok": True, "data": None})
    except Exception as e:
        logger.error(f"Get contractor payments failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/contractor-payments", methods=["POST"])
def save_contractor_payments():
    """Save contractor payments for a sheet."""
    body = request.get_json(force=True)
    sheet = body.get("sheet", "")
    data = body.get("data", [])
    if not sheet:
        return jsonify({"ok": False, "error": "Missing sheet"}), 400
    try:
        delivery_notes_db.save_contractor_payments(sheet, data)
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"Save contractor payments failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _fetch_trial_map_from_priority(url, hdrs, auth):
    """Fetch all ACCOUNTS from Priority and build trial balance map."""
    trial_map = {}
    skip = 0
    while True:
        acc_url = f"{url}/ACCOUNTS?$select=ACCNAME,TRIALBALCODE,TRIALBALDES&$top=500&$skip={skip}"
        r = http_requests.get(acc_url, headers=hdrs, auth=auth, timeout=60)
        if r.status_code >= 400:
            break
        rows = r.json().get("value", [])
        if not rows:
            break
        for a in rows:
            an = (a.get("ACCNAME") or "").strip()
            if an:
                trial_map[an] = {
                    "code": (a.get("TRIALBALCODE") or "").strip(),
                    "desc": (a.get("TRIALBALDES") or "").strip(),
                }
        skip += len(rows)
        if len(rows) < 500:
            break
        if skip > 50000:
            break
    return trial_map


def _get_trial_map(url, hdrs, auth, force_refresh=False):
    """Return trial balance map from DB cache. Only refreshes from Priority on explicit force_refresh."""
    if force_refresh:
        trial_map = _fetch_trial_map_from_priority(url, hdrs, auth)
        if trial_map:
            try:
                delivery_notes_db.save_accounts_cache(trial_map)
            except Exception as e:
                logger.error(f"Failed to save accounts cache: {e}")
        return trial_map

    # Always read from DB - never auto-fetch
    cached = delivery_notes_db.get_accounts_cache()
    if cached and cached.get("data"):
        return cached["data"]
    return {}


@app.route("/api/reports/accounts-status", methods=["GET"])
def accounts_status():
    """Get last sync info for accounts cache."""
    try:
        cached = delivery_notes_db.get_accounts_cache()
        if not cached:
            return jsonify({"ok": True, "count": 0, "updatedAt": None})
        return jsonify({
            "ok": True,
            "count": cached.get("count", 0),
            "updatedAt": cached.get("updated_at", ""),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/reports/sync-accounts", methods=["POST"])
def sync_accounts():
    """Force refresh accounts trial balance map from Priority."""
    try:
        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        hdrs = {"Accept": "application/json", "OData-Version": "4.0"}
        trial_map = _get_trial_map(url, hdrs, auth, force_refresh=True)
        return jsonify({"ok": True, "count": len(trial_map)})
    except Exception as e:
        logger.error(f"Sync accounts failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/reports/profit-loss", methods=["POST"])
def report_profit_loss():
    """Fetch FNCTRANS for accounts matching '4XX-{branch}' in given date range."""
    try:
        body = request.get_json(force=True)
        branch = str(body.get("branch", "")).strip()
        date_type = body.get("dateType", "FNCDATE")  # FNCDATE or BALDATE
        date_from = body.get("dateFrom", "")
        date_to = body.get("dateTo", "")

        if not branch or not date_from or not date_to:
            return jsonify({"ok": False, "error": "Missing branch or dates"}), 400
        if date_type not in ("FNCDATE", "BALDATE"):
            date_type = "FNCDATE"

        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        hdrs = {"Accept": "application/json", "OData-Version": "4.0"}

        # Fetch finalized transactions for this branch in date range
        # OData filter
        flt = (
            f"FINAL eq 'Y' and BRANCHNAME eq '{branch}' and "
            f"{date_type} ge {date_from}T00:00:00Z and {date_type} le {date_to}T23:59:59Z"
        )
        select = "FNCNUM,FNCDATE,BALDATE,CURDATE,DETAILS,REFERENCE"
        expand = "FNCITEMS_SUBFORM($select=ACCNAME,ACCDES,DEBIT1,CREDIT1,IACCNAME,DETAILS)"

        all_rows = []
        skip = 0
        while True:
            api_url = f"{url}/FNCTRANS?$filter={flt}&$select={select}&$expand={expand}&$top=500&$skip={skip}&$orderby={date_type} desc"
            resp = http_requests.get(api_url, headers=hdrs, auth=auth, timeout=60)
            if resp.status_code >= 400:
                logger.error(f"P&L query failed: {resp.status_code} {resp.text[:200]}")
                return jsonify({"ok": False, "error": resp.text[:200]}), 500
            rows = resp.json().get("value", [])
            if not rows:
                break
            all_rows.extend(rows)
            skip += len(rows)
            if len(rows) < 500:
                break
            if skip > 50000:
                break

        # Categorize by TRIALBALCODE: 400-490 = תקבולים, 6XX = הוצאות, 250-259 = הלוואות, else אחר
        def categorize(acc):
            t = trial_map.get((acc or "").strip())
            if not t:
                return 'אחר'
            code = (t.get('code') or '').strip()
            try:
                n = int(code)
                if 400 <= n <= 490:
                    return 'תקבולים'
                if 250 <= n <= 259:
                    return 'הלוואות'
                if n == 163:
                    return 'חברות קשורות'
            except ValueError:
                pass
            if code.startswith('6'):
                return 'הוצאות'
            return 'אחר'

        # Get cached trial balance map (fetched once per hour)
        trial_map = _get_trial_map(url, hdrs, auth)

        def trial_section(acc):
            t = trial_map.get((acc or "").strip())
            if not t:
                return ''
            code = t.get('code', '')
            desc = t.get('desc', '')
            if code and desc:
                return f"{code} - {desc}"
            return code or desc or ''

        def trial_code(acc):
            t = trial_map.get((acc or "").strip())
            return (t.get('code') if t else '') or ''

        result_rows = []
        for r in all_rows:
            # Skip P-prefix transactions (חשבוניות עסקה - not real revenue/expense)
            fncnum = (r.get('FNCNUM') or '').strip()
            if fncnum.startswith('P'):
                continue

            items = r.get('FNCITEMS_SUBFORM') or []
            for item in items:
                acc = (item.get('ACCNAME') or '').strip()
                if not acc:
                    continue
                acc_desc = item.get('ACCDES') or ''
                opp = (item.get('IACCNAME') or '').strip()
                debit = float(item.get('DEBIT1') or 0)
                credit = float(item.get('CREDIT1') or 0)
                line_details = item.get('DETAILS') or r.get('DETAILS', '')

                result_rows.append({
                    "category": categorize(acc),
                    "trialSection": trial_section(acc),
                    "trialCode": trial_code(acc),
                    "account": acc,
                    "accountDesc": acc_desc,
                    "fncnum": fncnum,
                    "fncDate": r.get('FNCDATE', ''),
                    "balDate": r.get('BALDATE', ''),
                    "details": line_details,
                    "oppAccount": opp,
                    "oppAccountDesc": '',
                    "debit": debit,
                    "credit": credit,
                    "reference": r.get('REFERENCE', ''),
                })

        return jsonify({"ok": True, "rows": result_rows, "totalScanned": len(all_rows)})
    except Exception as e:
        logger.error(f"P&L report failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/energy/charging-months", methods=["GET"])
def list_charging_months():
    """List all months that have saved charging sessions."""
    try:
        months = delivery_notes_db.list_charging_months()
        return jsonify({"ok": True, "months": months})
    except Exception as e:
        logger.error(f"List charging months failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/energy/charging-sessions", methods=["GET"])
def get_charging_sessions():
    """Get saved charging sessions for a specific month."""
    try:
        month = request.args.get("month", "").strip()
        if not month:
            return jsonify({"ok": False, "error": "Missing month parameter"}), 400
        result = delivery_notes_db.get_charging_sessions(month)
        if not result:
            return jsonify({"ok": True, "rows": [], "count": 0, "fileName": "", "updatedAt": "", "month": month})
        return jsonify({
            "ok": True,
            "rows": result.get("rows", []),
            "count": result.get("count", 0),
            "fileName": result.get("file_name", ""),
            "updatedAt": result.get("updated_at", ""),
            "month": result.get("month", month),
        })
    except Exception as e:
        logger.error(f"Get charging sessions failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/energy/charging-sessions", methods=["POST"])
def save_charging_sessions():
    """Save charging sessions to DB for a specific month."""
    try:
        body = request.get_json(force=True)
        month = (body.get("month") or "").strip()
        rows = body.get("rows", [])
        file_name = body.get("fileName", "")
        if not month:
            return jsonify({"ok": False, "error": "Missing month"}), 400
        if not rows:
            return jsonify({"ok": False, "error": "No rows"}), 400
        delivery_notes_db.save_charging_sessions(month, rows, file_name)
        return jsonify({"ok": True, "count": len(rows), "month": month})
    except Exception as e:
        logger.error(f"Save charging sessions failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


_HEBREW_MONTHS = {
    1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל",
    5: "מאי", 6: "יוני", 7: "יולי", 8: "אוגוסט",
    9: "ספטמבר", 10: "אוקטובר", 11: "נובמבר", 12: "דצמבר",
}


@app.route("/api/energy/create-invoices", methods=["POST"])
def energy_create_invoices():
    """Create draft customer invoices in Priority for energy charging.

    Body:
      month: "3.26"
      customers: list of { custname, total } - one per customer
      limit: optional, max number to create (for testing)
    """
    try:
        body = request.get_json(force=True)
        month_str = (body.get("month") or "").strip()
        customers = body.get("customers") or []
        limit = int(body.get("limit") or 0)

        if not month_str or not customers:
            return jsonify({"ok": False, "error": "Missing month or customers"}), 400

        # Parse month "3.26" → month=3, year=2026
        m_match = month_str.split(".")
        if len(m_match) != 2:
            return jsonify({"ok": False, "error": "Invalid month format. Expected M.YY"}), 400
        try:
            mm = int(m_match[0])
            yy = int(m_match[1])
            year = 2000 + yy if yy < 100 else yy
        except ValueError:
            return jsonify({"ok": False, "error": "Invalid month numbers"}), 400

        # Invoice date = first of NEXT month
        if mm == 12:
            inv_year = year + 1
            inv_month = 1
        else:
            inv_year = year
            inv_month = mm + 1
        inv_date = f"{inv_year:04d}-{inv_month:02d}-01"

        # Hebrew month name for the file's month
        hebrew_month = _HEBREW_MONTHS.get(mm, str(mm))
        details = f"עמלת גביה {hebrew_month} {year}"
        pdes = details  # also use as PDES

        # Limit if requested
        if limit > 0:
            customers = customers[:limit]

        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        hdrs = {"Accept": "application/json", "OData-Version": "4.0", "Content-Type": "application/json"}

        results = []
        for cust in customers:
            custname = str(cust.get("custname") or "").strip()
            total = float(cust.get("total") or 0)
            if not custname or total <= 0:
                results.append({"custname": custname, "ok": False, "error": "Missing custname or total"})
                continue

            try:
                # Net price (Priority adds VAT) — divide gross by 1.18
                net_price = round(total / 1.18, 2)
                body = {
                    "CUSTNAME": custname,
                    "BRANCHNAME": "110",
                    "IVDATE": inv_date,
                    "DETAILS": details,
                    "CINVOICEITEMS_SUBFORM": [{
                        "PARTNAME": "000",
                        "TQUANT": 1,
                        "PRICE": net_price,
                        "PDES": pdes,
                    }],
                }
                resp = http_requests.post(f"{url}/CINVOICES", json=body, headers=hdrs, auth=auth, timeout=30)
                if resp.status_code >= 400:
                    err_text = resp.text[:300]
                    results.append({"custname": custname, "ok": False, "error": f"Priority {resp.status_code}: {err_text}"})
                    continue
                data = resp.json()
                ivnum = data.get("IVNUM", "")
                results.append({
                    "custname": custname,
                    "ok": True,
                    "ivnum": ivnum,
                    "amount": total,
                    "netPrice": net_price,
                })
            except Exception as e:
                results.append({"custname": custname, "ok": False, "error": str(e)})

        return jsonify({"ok": True, "invDate": inv_date, "details": details, "results": results})
    except Exception as e:
        logger.error(f"Energy create invoices failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/energy/sync-customers", methods=["POST"])
def sync_customers_phone():
    """Force fetch all customers from Priority and save phone-map to DB."""
    try:
        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        hdrs = {"Accept": "application/json", "OData-Version": "4.0"}

        phone_map = {}  # normalized phone → {custname, custdes}
        skip = 0
        while True:
            api_url = f"{url}/CUSTOMERS?$select=CUSTNAME,CUSTDES,PHONE&$top=500&$skip={skip}"
            resp = http_requests.get(api_url, headers=hdrs, auth=auth, timeout=60)
            resp.raise_for_status()
            rows = resp.json().get("value", [])
            if not rows:
                break
            for r in rows:
                p = (r.get("PHONE") or "").strip()
                if not p:
                    continue
                digits = "".join(c for c in p if c.isdigit())
                if not digits:
                    continue
                key = digits[-9:] if len(digits) >= 9 else digits
                if key not in phone_map:
                    phone_map[key] = {"custname": r.get("CUSTNAME", ""), "custdes": r.get("CUSTDES", "")}
            skip += len(rows)
            if len(rows) < 500:
                break

        delivery_notes_db.save_customers_phone_cache(phone_map)
        return jsonify({"ok": True, "count": len(phone_map)})
    except Exception as e:
        logger.error(f"Sync customers failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/energy/customers-status", methods=["GET"])
def customers_phone_status():
    """Get last sync info for customers cache."""
    try:
        cached = delivery_notes_db.get_customers_phone_cache()
        if not cached:
            return jsonify({"ok": True, "count": 0, "updatedAt": None})
        return jsonify({"ok": True, "count": cached.get("count", 0), "updatedAt": cached.get("updated_at", "")})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/energy/customers-by-phone", methods=["POST"])
def customers_by_phone():
    """Look up customer numbers by phones from DB cache (NOT from Priority)."""
    try:
        body = request.get_json(force=True)
        phones = body.get("phones", [])
        if not phones:
            return jsonify({"ok": True, "results": {}})

        cached = delivery_notes_db.get_customers_phone_cache()
        phone_map = (cached or {}).get("data", {})
        if not phone_map:
            return jsonify({"ok": False, "error": "אין נתוני לקוחות ב-DB. סנכרן קודם."}), 400

        results = {}
        for phone in phones:
            digits = "".join(c for c in str(phone) if c.isdigit())
            key = digits[-9:] if len(digits) >= 9 else digits
            if key in phone_map:
                results[phone] = phone_map[key]

        return jsonify({"ok": True, "results": results, "totalScanned": len(phone_map)})
    except Exception as e:
        logger.error(f"Customers by phone lookup failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/contractor-lookup", methods=["POST"])
def contractor_lookup():
    """Look up contractor details from Priority SUPPLIERS by account number (SUPNAME)."""
    try:
        body = request.get_json(force=True)
        accounts = body.get("accounts", [])  # list of Priority SUPNAME values
        if not accounts:
            return jsonify({"ok": True, "results": {}})

        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        hdrs = {"Accept": "application/json", "OData-Version": "4.0"}

        results = {}
        for acc in accounts:
            acc = str(acc).strip()
            if not acc:
                continue
            api_url = f"{url}/SUPPLIERS?$filter=SUPNAME eq '{acc}'&$select=SUPNAME,SUPDES,VATNUM&$top=1"
            resp = http_requests.get(api_url, headers=hdrs, auth=auth, timeout=15)
            if resp.status_code < 400:
                rows = resp.json().get("value", [])
                if rows:
                    results[acc] = {
                        "supdes": rows[0].get("SUPDES", ""),
                        "vatnum": rows[0].get("VATNUM", ""),
                    }

        return jsonify({"ok": True, "results": results})
    except Exception as e:
        logger.error(f"Contractor lookup failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Cinvoices (חשבוניות מרכזות) ─────────────────────────────

@app.route("/api/hr/cinvoice", methods=["POST"])
def create_cinvoice():
    """Save a cinvoice draft to DB."""
    try:
        data = request.get_json(force=True)
        customer_num = (data.get("customerNum") or "").strip()
        if customer_num.endswith("-102"):
            customer_num = customer_num[:-4]
        if not customer_num:
            return jsonify({"ok": False, "error": "Missing customerNum"}), 400

        raw_items = data.get("items", [])
        if not raw_items:
            return jsonify({"ok": False, "error": "Missing items"}), 400

        db_items = []
        for item in raw_items:
            db_items.append({
                "partname": str(item.get("profNum", "")).strip(),
                "pdes": str(item.get("profName", "")).strip(),
                "tquant": item.get("hours", 0),
                "price": item.get("rate", 0),
            })

        result = delivery_notes_db.save_delivery_note(
            customer_num=customer_num,
            customer_name=data.get("customerName", ""),
            site_name=data.get("siteName", ""),
            details=data.get("details", ""),
            items=db_items,
        )
        # Tag it as cinvoice type
        delivery_notes_db.update_delivery_note(result["id"], {"doc_type": "cinvoice"})

        return jsonify({"ok": True, "id": result["id"], "status": "draft"})
    except Exception as e:
        logger.error(f"Cinvoice save failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/cinvoices/<note_id>", methods=["GET"])
def get_cinvoice(note_id):
    """Get a single cinvoice."""
    try:
        note = delivery_notes_db.get_delivery_note(note_id)
        if not note:
            return jsonify({"ok": False, "error": "Not found"}), 404
        return jsonify({"ok": True, "note": note})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/cinvoices/<note_id>", methods=["PUT"])
def update_cinvoice(note_id):
    """Update a cinvoice draft."""
    try:
        data = request.get_json(force=True)
        note = delivery_notes_db.get_delivery_note(note_id)
        if not note:
            return jsonify({"ok": False, "error": "Not found"}), 404
        if note.get("status") == "sent":
            return jsonify({"ok": False, "error": "Cannot edit a sent cinvoice"}), 400
        updates = {}
        for field in ("details", "items", "customer_num", "customer_name", "site_name"):
            if field in data:
                updates[field] = data[field]
        if updates:
            delivery_notes_db.update_delivery_note(note_id, updates)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/cinvoices/<note_id>", methods=["DELETE"])
def delete_cinvoice(note_id):
    """Delete a cinvoice draft."""
    try:
        note = delivery_notes_db.get_delivery_note(note_id)
        if not note:
            return jsonify({"ok": False, "error": "Not found"}), 404
        if note.get("status") == "sent":
            return jsonify({"ok": False, "error": "Cannot delete a sent cinvoice"}), 400
        delivery_notes_db.delete_delivery_note(note_id)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/cinvoices/<note_id>/send", methods=["POST"])
def send_cinvoice_to_priority(note_id):
    """Send a cinvoice draft to Priority (CINVOICES)."""
    try:
        note = delivery_notes_db.get_delivery_note(note_id)
        if not note:
            return jsonify({"ok": False, "error": "Not found"}), 404
        if note.get("status") == "sent":
            return jsonify({"ok": False, "error": "Already sent", "ivnum": note.get("docno")}), 400

        customer_num = note["customer_num"]
        items = note.get("items", [])

        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
            "OData-Version": "4.0",
        }

        subform_items = []
        for item in items:
            subform_items.append({
                "PARTNAME": str(item.get("partname", "")).strip(),
                "PDES": str(item.get("pdes", "")).strip(),
                "TQUANT": item.get("tquant", 0),
                "PRICE": item.get("price", 0),
            })

        # Look up site code (DCODE) from sites cache
        site_code = ""
        site_name = note.get("site_name", "")
        if site_name:
            cached_sites = delivery_notes_db.load_sites_cache()
            if cached_sites:
                for s in cached_sites.get("sites", []):
                    if s.get("name") == site_name and s.get("custCode") == customer_num:
                        site_code = s.get("code", "")
                        break

        body = {
            "CUSTNAME": customer_num,
            "BRANCHNAME": "102",
            "DETAILS": note.get("details", ""),
            "CINVOICEITEMS_SUBFORM": subform_items,
        }
        if site_code:
            body["DCODE"] = site_code

        logger.info(f"[cinvoice] Sending {note_id} to Priority for customer {customer_num}, site {site_code}")
        resp = http_requests.post(f"{url}/CINVOICES", json=body, headers=headers, auth=auth, timeout=30)
        if not resp.ok:
            error_text = resp.text[:500]
            logger.error(f"[cinvoice] Priority error {resp.status_code}: {error_text}")
            delivery_notes_db.mark_error(note_id, f"Priority {resp.status_code}: {error_text}")
            return jsonify({"ok": False, "error": f"Priority error {resp.status_code}: {error_text}"}), 500

        result = resp.json()
        ivnum = result.get("IVNUM", "")
        delivery_notes_db.mark_sent(note_id, ivnum)
        logger.info(f"[cinvoice] Sent {note_id} → {ivnum}")

        return jsonify({"ok": True, "ivnum": ivnum})
    except Exception as e:
        logger.error(f"Send cinvoice failed: {e}")
        delivery_notes_db.mark_error(note_id, str(e))
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/delivery-note", methods=["POST"])
def create_delivery_note():
    """Save a delivery note to DB (draft). Does NOT send to Priority yet."""
    try:
        data = request.get_json(force=True)
        customer_num = (data.get("customerNum") or "").strip()
        # Strip branch suffix (e.g. "50254-102" → "50254")
        if customer_num.endswith("-102"):
            customer_num = customer_num[:-4]
        if not customer_num:
            return jsonify({"ok": False, "error": "Missing customerNum"}), 400

        raw_items = data.get("items", [])
        if not raw_items:
            return jsonify({"ok": False, "error": "Missing items"}), 400

        customer_name = data.get("customerName", "")
        site_name = data.get("siteName", "")
        details = data.get("details", site_name)

        # Normalize items for DB
        db_items = []
        for item in raw_items:
            db_items.append({
                "partname": str(item.get("profNum", "")).strip(),
                "pdes": str(item.get("profName", "")).strip(),
                "tquant": item.get("hours", 0),
                "price": item.get("rate", 0),
            })

        result = delivery_notes_db.save_delivery_note(
            customer_num=customer_num,
            customer_name=customer_name,
            site_name=site_name,
            details=details,
            items=db_items,
        )
        logger.info(f"[delivery-note] Saved draft {result['id']} for customer {customer_num}")

        return jsonify({"ok": True, "id": result["id"], "status": "draft"})
    except Exception as e:
        logger.error(f"Delivery note save failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/delivery-notes", methods=["GET"])
def list_delivery_notes():
    """List delivery notes, optionally filtered by status."""
    try:
        status = request.args.get("status")
        notes = delivery_notes_db.list_delivery_notes(status=status)
        return jsonify({"ok": True, "notes": notes})
    except Exception as e:
        logger.error(f"List delivery notes failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/delivery-notes/<note_id>", methods=["GET"])
def get_delivery_note(note_id):
    """Get a single delivery note."""
    try:
        note = delivery_notes_db.get_delivery_note(note_id)
        if not note:
            return jsonify({"ok": False, "error": "Not found"}), 404
        return jsonify({"ok": True, "note": note})
    except Exception as e:
        logger.error(f"Get delivery note failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/delivery-notes/<note_id>", methods=["PUT"])
def update_delivery_note_route(note_id):
    """Update a delivery note (items, details, etc.)."""
    try:
        data = request.get_json(force=True)
        note = delivery_notes_db.get_delivery_note(note_id)
        if not note:
            return jsonify({"ok": False, "error": "Not found"}), 404
        if note.get("status") == "sent":
            return jsonify({"ok": False, "error": "Cannot edit a sent delivery note"}), 400

        updates = {}
        for field in ("details", "items", "customer_num", "customer_name", "site_name"):
            if field in data:
                updates[field] = data[field]

        if updates:
            delivery_notes_db.update_delivery_note(note_id, updates)

        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"Update delivery note failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/delivery-notes/<note_id>", methods=["DELETE"])
def delete_delivery_note_route(note_id):
    """Delete a delivery note (only drafts)."""
    try:
        note = delivery_notes_db.get_delivery_note(note_id)
        if not note:
            return jsonify({"ok": False, "error": "Not found"}), 404
        if note.get("status") == "sent":
            return jsonify({"ok": False, "error": "Cannot delete a sent delivery note"}), 400

        delivery_notes_db.delete_delivery_note(note_id)
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"Delete delivery note failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/hr/delivery-notes/<note_id>/send", methods=["POST"])
def send_delivery_note_to_priority(note_id):
    """Send a draft delivery note to Priority (DOCUMENTS_D)."""
    try:
        note = delivery_notes_db.get_delivery_note(note_id)
        if not note:
            return jsonify({"ok": False, "error": "Not found"}), 404
        if note.get("status") == "sent":
            return jsonify({"ok": False, "error": "Already sent", "docno": note.get("docno")}), 400

        customer_num = note["customer_num"]
        items = note.get("items", [])

        url = PRIORITY_URL_REAL
        auth = HTTPBasicAuth(
            os.getenv("PRIORITY_USERNAME", ""),
            os.getenv("PRIORITY_PASSWORD", ""),
        )
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
            "OData-Version": "4.0",
        }

        subform_items = []
        for item in items:
            subform_items.append({
                "PARTNAME": str(item.get("partname", "")).strip(),
                "PDES": str(item.get("pdes", "")).strip(),
                "TQUANT": item.get("tquant", 0),
                "PRICE": item.get("price", 0),
            })

        # Look up site code (DCODE) from sites cache
        site_code = ""
        site_name = note.get("site_name", "")
        if site_name:
            cached_sites = delivery_notes_db.load_sites_cache()
            if cached_sites:
                for s in cached_sites.get("sites", []):
                    if s.get("name") == site_name and s.get("custCode") == customer_num:
                        site_code = s.get("code", "")
                        break

        body = {
            "CUSTNAME": customer_num,
            "BRANCHNAME": "102",
            "DETAILS": note.get("details", ""),
            "TRANSORDER_D_SUBFORM": subform_items,
        }
        if site_code:
            body["DCODE"] = site_code

        logger.info(f"[delivery-note] Sending {note_id} to Priority for customer {customer_num}, site {site_code}")
        resp = http_requests.post(f"{url}/DOCUMENTS_D", json=body, headers=headers, auth=auth, timeout=30)
        if not resp.ok:
            error_text = resp.text[:500]
            logger.error(f"[delivery-note] Priority error {resp.status_code}: {error_text}")
            delivery_notes_db.mark_error(note_id, f"Priority {resp.status_code}: {error_text}")
            return jsonify({"ok": False, "error": f"Priority error {resp.status_code}: {error_text}"}), 500

        result = resp.json()
        docno = result.get("DOCNO", result.get("DOCNUM", ""))
        delivery_notes_db.mark_sent(note_id, docno)
        logger.info(f"[delivery-note] Sent {note_id} → {docno}")

        return jsonify({"ok": True, "docno": docno})
    except Exception as e:
        logger.error(f"Send delivery note failed: {e}")
        delivery_notes_db.mark_error(note_id, str(e))
        return jsonify({"ok": False, "error": str(e)}), 500


if __name__ == "__main__":
    print("Urban Group Backend API")
    print(f"Priority Demo: {PRIORITY_URL_DEMO}")
    print(f"Priority Real: {PRIORITY_URL_REAL}")
    print("Starting on http://localhost:5000")
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True,
        exclude_patterns=["*/WindowsApps/*", "*/encodings/*"],
    )
