"""
1000-Invoice Batch Flow
Reads invoices from Excel file and creates each one in Priority
by calling agent 200's create_invoice function.
"""

import sys
import os
import io
import json
import importlib.util
from pathlib import Path
from datetime import datetime

import openpyxl

# Fix Windows console encoding for Hebrew
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# Project root
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent

# Load agent 200 module (uses importlib because folder name has hyphens)
# Agent 200's module-level code sets sys.stdout to UTF-8
agent_200_path = PROJECT_ROOT / "agents" / "200-invoices" / "200-invoice_writer.py"
spec = importlib.util.spec_from_file_location("invoice_writer", agent_200_path)
invoice_writer = importlib.util.module_from_spec(spec)
sys.modules["invoice_writer"] = invoice_writer
spec.loader.exec_module(invoice_writer)

# Load agent 100 module (customer reader)
agent_100_path = PROJECT_ROOT / "agents" / "100-customer" / "100-customer_reader.py"
spec_100 = importlib.util.spec_from_file_location("customer_reader", agent_100_path)
customer_reader = importlib.util.module_from_spec(spec_100)
sys.modules["customer_reader"] = customer_reader
spec_100.loader.exec_module(customer_reader)

# Input / Output paths
INPUT_FILE = PROJECT_ROOT / "input" / "חשבונית עמלת גבייה - טמפלט.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "output"

# Max invoices to process (set to None for all)
MAX_INVOICES = 4


def load_invoices_from_excel():
    """Load invoice data from Excel file.

    Excel structure (row 2 = headers, row 3+ = data):
      B: מס (row number)
      C: תאריך חשבונית
      D: פרטים (DETAILS)
      E: סניף
      F: מספר לקוח פריוריטי (CUSTNAME)
      G: שם לקוח
      H: מקט (PARTNAME)
      I: תאור מוצר (PDES)
      J: כמות (TQUANT)
      K: סכום לפני מעמ (PRICE)
      L: סכום כולל מעמ
    """
    if not INPUT_FILE.exists():
        print(f"Error: Input file not found: {INPUT_FILE}")
        sys.exit(1)

    # Load with data_only=True to get calculated values from formulas
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb.active

    invoices = []

    # Data starts at row 3 (row 1 = empty, row 2 = headers)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        row_num = row[1].value       # B: מס
        date_val = row[2].value      # C: תאריך חשבונית
        details = row[3].value       # D: פרטים
        branch = row[4].value        # E: סניף
        custname = row[5].value      # F: מספר לקוח פריוריטי
        cust_label = row[6].value    # G: שם לקוח
        partname = row[7].value      # H: מקט
        part_desc = row[8].value     # I: תאור מוצר
        quantity = row[9].value      # J: כמות
        price_no_vat = row[10].value # K: סכום לפני מעמ
        price_with_vat = row[11].value  # L: סכום כולל מעמ

        # Skip empty rows
        if not custname:
            continue

        # Format date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        elif date_val:
            date_str = str(date_val)
        else:
            date_str = datetime.now().strftime("%Y-%m-%d")

        # Always calculate price before VAT from column L (with VAT)
        # Column K may contain stale formula cache values (data_only=True)
        if price_with_vat is not None:
            price_no_vat = round(price_with_vat / 1.18, 2)

        item = {
            "PARTNAME": str(partname).strip(),
            "TQUANT": quantity or 1,
            "PRICE": price_no_vat or 0,
        }

        if part_desc:
            item["PDES"] = str(part_desc).strip()

        invoices.append({
            "row": row_num,
            "CUSTNAME": str(custname).strip(),
            "CUST_LABEL": str(cust_label or "").strip(),
            "IVDATE": date_str,
            "BRANCHNAME": str(branch or "000").strip(),
            "DETAILS": str(details or "").strip(),
            "items": [item],
        })

    wb.close()
    return invoices


def main():
    print("=" * 60)
    print("  1000-Invoice Batch Flow - Priority Cloud")
    print("  קריאה מקובץ Excel")
    print("=" * 60)
    print()

    invoice_writer.validate_config()

    print(f"Connecting to: {invoice_writer.PRIORITY_URL}")
    print(f"User: {invoice_writer.PRIORITY_USERNAME}")
    print()

    # Fetch customer list from Priority for validation
    print("Fetching customer list from Priority...")
    customer_map = {}
    try:
        customers = customer_reader.fetch_customers(top=9999)
        customer_map = {c["CUSTNAME"]: c["CUSTDES"] for c in customers}
        print(f"Loaded {len(customer_map)} customers from Priority")
    except Exception as e:
        print(f"Warning: Could not fetch customer list: {e}")
    print()

    all_invoices = load_invoices_from_excel()
    print(f"Loaded {len(all_invoices)} invoices from: {INPUT_FILE.name}")

    # Limit to MAX_INVOICES
    invoices = all_invoices[:MAX_INVOICES] if MAX_INVOICES else all_invoices
    print(f"Processing {len(invoices)} invoices (limit: {MAX_INVOICES or 'none'})")
    print()

    results = []

    for i, inv in enumerate(invoices, 1):
        customer = inv["CUSTNAME"]
        cust_label = inv["CUST_LABEL"]
        date = inv["IVDATE"]
        branch = inv["BRANCHNAME"]
        details = inv["DETAILS"]
        items = inv["items"]

        priority_name = customer_map.get(customer)

        print(f"--- Invoice {i}/{len(invoices)} (row {inv['row']}) ---")
        print(f"Customer: {customer} ({cust_label})")
        if priority_name:
            print(f"Priority name: {priority_name}")
        elif customer_map:
            print(f"WARNING: Customer {customer} not found in Priority!")
        print(f"Date: {date}, Branch: {branch}")
        print(f"Details: {details}")
        pdes = items[0].get('PDES', '')
        print(f"Items: Part={items[0]['PARTNAME']}, Desc={pdes}, Qty={items[0]['TQUANT']}, Price={items[0]['PRICE']}")

        try:
            result = invoice_writer.create_invoice(customer, date, branch, items, details=details)
            ivnum = result.get("IVNUM", "N/A")
            totprice = result.get("TOTPRICE", 0)
            print(f"OK - Invoice {ivnum} created, Total: {totprice}")
            results.append({
                "index": i,
                "row": inv["row"],
                "customer": customer,
                "customer_name": cust_label,
                "status": "OK",
                "ivnum": ivnum,
                "totprice": totprice,
            })
        except Exception as e:
            error_msg = str(e)
            print(f"FAILED - {error_msg}")
            results.append({
                "index": i,
                "row": inv["row"],
                "customer": customer,
                "customer_name": cust_label,
                "status": "FAILED",
                "error": error_msg,
            })

        print()

    # Summary
    success = sum(1 for r in results if r["status"] == "OK")
    failed = sum(1 for r in results if r["status"] == "FAILED")

    print("=" * 60)
    print(f"  Batch Complete: {success} OK, {failed} Failed, {len(results)} Total")
    print("=" * 60)

    # Save output
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "1000-invoice_batch.txt"

    with open(output_file, "w", encoding="utf-8") as f:
        f.write("1000-Invoice Batch Flow - Priority Cloud\n")
        f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source: {invoice_writer.PRIORITY_URL}\n")
        f.write(f"Input: {INPUT_FILE.name}\n")
        f.write(f"Limit: {MAX_INVOICES or 'all'}\n")
        f.write("\n")
        f.write(f"Results: {success} OK, {failed} Failed, {len(results)} Total\n")
        f.write("-" * 50 + "\n")

        for r in results:
            if r["status"] == "OK":
                f.write(f"  [row {r['row']}] {r['customer']} ({r['customer_name']}) -> "
                        f"Invoice: {r['ivnum']}, Total: {r['totprice']}\n")
            else:
                f.write(f"  [row {r['row']}] {r['customer']} ({r['customer_name']}) -> "
                        f"FAILED: {r['error']}\n")

        f.write("-" * 50 + "\n")

    print()
    print(f"Saved to: {output_file}")


if __name__ == "__main__":
    main()
